# -*- coding: utf-8 -*-
"""
worker_per_process.py

일일생산계획표(.XLSX) 파일들을 모두 읽어서, 각 날짜 생산실적 항목의 '리웍'
텍스트를 라인/공정 종류('단발', '로타리', '셀라인', 'SET', '접지', '충진',
'포장', 기타)로 분류하고, 분류별로 (생산수량 / 인원), (계획수량 / 인원)
두 가지 비율의 분포를 계산 및 시각화한다. ('수축'은 셀라인 공정의 연장선
작업으로 보고 별도 카테고리를 두지 않고 '셀라인'에 합쳐서 분류한다.)
또한 제품군(마스크/튜브/용기 등)별로 어떤 리웍이 주로 쓰이는지도 함께
집계한다.

목적: 같은 라인(공정)인데도 날짜/건에 따라 생산수량 대비 투입 인원이
얼마나 들쭉날쭉한지 확인해서, 인원이 과다 투입된 날(=1인당 생산량이
비정상적으로 낮은 케이스)을 찾아내기 위함. 생산수량(실제 생산한 양) 기준과
계획수량(원래 계획했던 양) 기준을 나란히 비교하면, "계획 대비로는 정상인데
실제 생산이 저조해서 비율이 낮아진 건지" 아니면 "애초에 계획 단계부터
인원 대비 계획량이 적었던 건지"를 구분해서 볼 수 있다.

엑셀 구조 메모 (사람이 나중에 참고할 수 있도록):
  - 파일 하나에 여러 시트(예: '2025년 1월' ~ '2025년 7월')가 있고, 시트별로
    같은 마스터 생산계획표가 '해당 시점까지의 스냅샷'으로 저장되어 있음.
  - 2행: 제품군/설비/발주처/품번/품명/단위 등 기본 정보 헤더 + 월 헤더(병합)
  - 3행: 일자 헤더('3일', '4일', ...) (병합)
  - 4행: 각 날짜 블록의 하위 헤더('계획수량','생산수량','불량수량'(있는 시트만),
         '리웍','인원')가 5개(또는 4개) 컬럼 단위로 반복됨.
  - 시트마다 앞쪽 기본 정보 컬럼 구성이 다르고(예: 1~2월 시트는 '설비' 컬럼이
    있고 3월 이후 시트는 없음), 날짜 블록 폭도 4~5컬럼으로 다르므로 컬럼
    위치를 고정 인덱스로 가정하지 않고 매번 헤더 텍스트를 읽어서 동적으로
    찾는다.
  - '리웍' 컬럼은 숫자가 아니라 그날 작업 내용을 적은 자유 텍스트다.
    예: 'set', '셀라인', '단발', '로타리', '접지, 잔업', 'set,3라인' 등.
    이 텍스트 안에 라인 키워드가 포함되어 있으면 그 라인으로 분류한다.
"""

from __future__ import annotations

import argparse
import glob
import io
import os
import re
import sys

# Windows 콘솔(cp949 등) 기본 인코딩에서는 한글 출력이 깨지므로, 표준 출력/
# 에러 스트림을 UTF-8로 강제로 다시 감싼다. (파일 저장은 to_csv에서 별도로
# utf-8-sig를 지정하므로 이 처리와 무관하게 항상 정상 저장된다.)
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from collections import Counter, OrderedDict
from dataclasses import dataclass

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

# 그래프에 한글이 깨지지 않도록 한글 폰트 지정 (Windows 기본 내장 폰트).
plt.rcParams["font.family"] = "Malgun Gothic"
plt.rcParams["axes.unicode_minus"] = False  # 한글 폰트에서 마이너스(-) 기호가 깨지는 것 방지

DEFAULT_INPUT_DIR = r"C:\Users\USER\Desktop\유빈_생산계획\일일생산계획_예시"
DEFAULT_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

# 일자 헤더('3일', '10일' 등)를 알아보기 위한 정규식. 이 패턴에 안 맞는 헤더
# (예: '전월', '생산' 같은 요약 블록 라벨)는 날짜 블록에서 제외된다.
DAY_PATTERN = re.compile(r"^\d+일$")

# 시트 2행에서 찾을 기본 정보 컬럼 이름들 (find_base_columns에서 사용).
BASE_HEADER_NAMES = ["제품군", "설비", "발주처", "품번", "품명", "단위"]
# 시트 4행에서 찾을, 날짜 블록 안에 반복되는 하위 헤더 이름들 (find_date_blocks에서 사용).
BLOCK_LABELS = ["계획수량", "생산수량", "불량수량", "리웍", "인원"]

# 분류 우선순위: 먼저 매칭되는 키워드로 분류한다.
# 앞쪽 4개(셀라인/단발/로타리/SET)는 생산 라인 종류, 뒤쪽 3개(접지/충진/포장)는
# 라인 지정 없이 공정만 적힌 경우를 위한 카테고리. 예: '접지, 잔업' -> 접지,
# 'set, 접지' -> 라인이 우선이므로 SET.
# '수축'은 별도 카테고리로 두지 않고 '셀라인' 키워드 목록에 포함시켰다.
# 셀라인 공정의 연장선(후처리) 작업이라서 실무적으로 같은 셀라인 작업으로
# 보는 게 맞다는 판단. 그래서 '수축'만 단독으로 적힌 건('수축', '수축,포장'
# 등)도 이제 '셀라인'으로 분류된다.
CATEGORY_KEYWORDS = OrderedDict(
    [
        ("셀라인", ["셀라인", "수축"]),
        ("단발", ["단발"]),
        ("로타리", ["로타리"]),
        ("SET", ["set"]),
        ("접지", ["접지"]),
        ("충진", ["충진"]),
        ("포장", ["포장"]),
    ]
)
CATEGORY_ORDER = list(CATEGORY_KEYWORDS.keys()) + ["기타"]


def normalize_header(v):
    """엑셀 헤더 셀 값에서 줄바꿈(\\n)을 제거해 비교하기 쉬운 문자열로 만든다.

    엑셀에 '계획\\n수량'처럼 셀 안에서 줄바꿈된 헤더가 많아서, 이 함수를 거치지
    않으면 '계획수량' 같은 코드 상의 상수 문자열과 절대 일치하지 않는다.
    문자열이 아닌 값(None, 숫자, 날짜 등)은 그대로 반환한다.
    """
    return v.replace("\n", "") if isinstance(v, str) else v


def classify_rework(text: str) -> str:
    """'리웍' 컬럼의 자유 텍스트 한 칸을 보고 라인/공정 카테고리 하나로 분류한다.

    CATEGORY_KEYWORDS를 위에서부터 순서대로 검사해서, 텍스트 안에 키워드가
    포함되어 있으면(부분 문자열 매치, 대소문자 무시) 그 카테고리를 반환한다.
    예) '셀라인,수축,잔업' -> 앞쪽에 '셀라인'이 있으므로 '셀라인'으로 분류.
        '접지, 잔업'       -> 라인 이름이 없고 '접지'만 있으므로 '접지'로 분류.
    아무 키워드에도 안 걸리면(두라인, 잔업, 재고메모 등 자잘한 표기들) '기타'로 묶는다.
    """
    t = str(text)
    for cat, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in t.lower():
                return cat
    return "기타"


def forward_fill_row(ws, row: int, max_col: int) -> dict:
    """지정한 행(row)을 왼쪽에서 오른쪽으로 훑으며, 비어있는(None) 칸을 바로
    직전에 나온 값으로 채워서 {열번호: 값} 딕셔너리로 반환한다.

    왜 필요한가: 엑셀에서 월 헤더('6월')나 일자 헤더('3일')는 여러 컬럼에
    걸쳐 셀 병합(merge)되어 있는데, openpyxl로 값을 읽으면 병합 범위의
    '맨 왼쪽 칸'에만 값이 들어있고 나머지 칸은 전부 None으로 읽힌다.
    예) 2행에서 Q2='6월'이고 R2~U2는 실제로는 병합되어 같은 '6월'이지만
        openpyxl은 R2~U2를 None으로 준다. 이 함수로 정방향 채우기(forward
        fill)를 하면 R2~U2도 '6월'로 채워져서, 날짜 블록이 시작하는 아무
        컬럼에서나 그 컬럼이 속한 월/일을 바로 조회할 수 있다.
    """
    vals = {}
    last = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None:
            last = v
        vals[c] = last
    return vals


def to_number(v):
    """엑셀 셀 값을 숫자(float)로 변환하되, 숫자가 아니면 None을 반환한다.

    '인원', '생산수량' 같은 칸에는 종종 숫자 대신 '예비라인', '~', '13명'처럼
    사람이 적어놓은 메모성 텍스트가 들어있다(엑셀 원본 확인 결과). 이런 값은
    분석에 쓸 수 없으므로 조용히 None으로 걸러내고, 호출부에서 None인 행은
    건너뛴다. bool은 int의 서브클래스라 isinstance(v, (int, float))에 True로
    걸리는 것을 막기 위해 별도로 먼저 체크한다.
    """
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


@dataclass
class Record:
    """엑셀에서 뽑아낸 '한 제품 x 하루' 단위의 생산 실적 한 건.

    예를 들어 '6월 3일에 이 품번을 셀라인에서 6명이 투입되어 4662개 생산'이
    한 Record가 된다. 이 Record들을 모아 DataFrame으로 만든 것이 이후 모든
    집계/그래프의 원재료(records.csv)가 된다.
    """

    source_file: str      # 어느 엑셀 파일에서 나왔는지 (디버깅/추적용)
    sheet: str             # 어느 시트('2025년 6월' 등)에서 나왔는지
    month: str             # 그 날짜 블록의 월 헤더 값 (예: '6월')
    day: str               # 그 날짜 블록의 일자 헤더 값 (예: '3일')
    product_group: str     # 제품군 (마스크/튜브/용기 등)
    product_code: str      # 품번
    product_name: str      # 품명
    plan_qty: float | None   # 그날 계획수량 (plan_ratio의 분자)
    prod_qty: float           # 그날 생산수량 (ratio의 분자)
    defect_qty: float | None  # 그날 불량수량 (있는 시트에서만 채워짐)
    rework_raw: str        # '리웍' 칸의 원본 텍스트 (예: '셀라인,수축,잔업')
    category: str          # rework_raw를 classify_rework()로 분류한 결과
    person: float              # 그날 투입 인원 (ratio/plan_ratio의 분모)
    ratio: float                # prod_qty / person = 1인당 생산수량 (실적 기준)
    plan_ratio: float | None    # plan_qty / person = 1인당 계획수량 (계획 기준).
                                 # plan_qty가 숫자가 아니면(빈칸, '-' 등) None.


def find_base_columns(ws, max_col: int) -> dict:
    """시트의 2행(헤더 행)을 보고 '제품군/설비/발주처/품번/품명/단위' 같은
    기본 정보 컬럼들이 각각 몇 번째 열에 있는지 찾아 {헤더이름: 열번호}로 반환한다.

    시트마다 이 기본 정보 컬럼의 개수/순서가 다르다(예: 1~2월 시트는 '설비'
    컬럼이 있지만 3월 이후 시트는 없어서 나머지 컬럼이 한 칸씩 왼쪽으로
    당겨져 있음). 그래서 열 번호를 코드에 고정으로 박아두지 않고, 매 시트마다
    헤더 텍스트로 실제 위치를 찾아낸다.
    """
    row2 = {c: normalize_header(ws.cell(row=2, column=c).value) for c in range(1, max_col + 1)}
    cols = {}
    for c, v in row2.items():
        if isinstance(v, str) and v.strip() in BASE_HEADER_NAMES and v.strip() not in cols:
            cols[v.strip()] = c
    return cols


def find_date_blocks(ws, max_col: int, row2_ff: dict, row3_ff: dict) -> list[tuple[str, str, dict]]:
    """시트의 4행(하위 헤더 행)을 스캔해서, 날짜별로 반복되는 컬럼 블록들을
    전부 찾아낸다. 각 블록은 (월, 일, {서브헤더이름: 열번호}) 형태다.

    동작 방식:
      1. 4행에서 값이 '계획수량'인 열들을 모두 찾는다 -> 각 열이 하나의
         날짜 블록의 시작 지점이다(한 블록은 계획수량/생산수량/불량수량(선택)
         /리웍/인원 순서로 최대 5칸을 차지한다).
      2. 각 블록 시작 열부터 오른쪽으로 최대 6칸까지 살펴보면서, 4행 값이
         BLOCK_LABELS(계획수량/생산수량/불량수량/리웍/인원) 중 하나면 그
         이름으로 열 번호를 기록한다. 다음 '계획수량'을 만나면 그 블록은
         끝난 것으로 보고 중단한다. (시트에 따라 '불량수량'이 없어서 블록
         폭이 4칸일 수도 있어 하드코딩된 5가 아니라 이렇게 유동적으로 찾는다.)
      3. forward_fill_row()로 미리 채워둔 2행(월)/3행(일) 값을 블록 시작
         열 기준으로 조회해서 그 블록이 몇 월 며칠 데이터인지 알아낸다.
      4. 일자 헤더가 '3일', '10일'처럼 숫자+'일' 형식인 블록만 실제 날짜
         데이터로 인정한다. (예: '전월'/'생산' 같은 요약 블록은 날짜 데이터가
         아니므로 DAY_PATTERN에 안 걸려서 자동으로 제외된다.)
    """
    row4 = {c: normalize_header(ws.cell(row=4, column=c).value) for c in range(1, max_col + 1)}
    block_starts = [c for c in range(1, max_col + 1) if row4.get(c) == "계획수량"]

    blocks = []
    for bs in block_starts:
        cols = {}
        c = bs
        limit = bs + 6
        while c < limit and c <= max_col:
            label = row4.get(c)
            if label == "계획수량" and c != bs:
                break
            if label in BLOCK_LABELS:
                cols[label] = c
            c += 1
        month = row2_ff.get(bs)
        day = row3_ff.get(bs)
        if isinstance(day, str) and DAY_PATTERN.match(day.strip()):
            blocks.append((str(month), day.strip(), cols))
    return blocks


def extract_records(file_path: str) -> list[Record]:
    """엑셀 파일 하나(모든 시트)를 읽어서 Record 리스트로 변환하는 메인 파서.

    시트 하나마다:
      1) find_base_columns()로 품번/품명 등 기본 정보 컬럼 위치를 찾는다.
         '품번'도 '품명'도 없는 시트는 우리가 기대하는 생산계획표 형식이
         아니라고 보고 건너뛴다.
      2) forward_fill_row()로 월/일 헤더를 채우고, find_date_blocks()로
         날짜별 컬럼 블록 목록을 얻는다.
      3) 데이터가 시작되는 5행부터 끝까지, 행(=제품 한 줄) x 블록(=하루)
         조합마다 리웍/인원/생산수량 값을 읽는다.
         - 리웍 칸이 비어있으면(그날 작업 안 함) 건너뛴다.
         - 인원이 숫자가 아니거나 0 이하이면 비율 계산이 불가능하므로 건너뛴다.
         - 생산수량이 숫자가 아니면(None 등) 건너뛴다. 단, 0은 유효한 값으로
           취급한다(인원은 투입됐는데 생산량이 0인 것도 의미 있는 관찰치이므로).
      4) 남은 유효한 조합만 classify_rework()로 카테고리를 매기고 ratio를
         계산해 Record로 쌓는다.
    """
    fname = os.path.basename(file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    records: list[Record] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column
        max_row = ws.max_row

        base_cols = find_base_columns(ws, max_col)
        if "품번" not in base_cols and "품명" not in base_cols:
            continue  # 예상한 표 형식이 아닌 시트는 건너뜀

        row2_ff = forward_fill_row(ws, 2, max_col)
        row3_ff = forward_fill_row(ws, 3, max_col)
        blocks = find_date_blocks(ws, max_col, row2_ff, row3_ff)
        if not blocks:
            continue

        pg_col = base_cols.get("제품군")
        code_col = base_cols.get("품번")
        name_col = base_cols.get("품명")

        for r in range(5, max_row + 1):
            product_group = ws.cell(row=r, column=pg_col).value if pg_col else None
            product_code = ws.cell(row=r, column=code_col).value if code_col else None
            product_name = ws.cell(row=r, column=name_col).value if name_col else None
            if product_group is None and product_code is None and product_name is None:
                continue

            for month, day, cols in blocks:
                if "리웍" not in cols or "인원" not in cols or "생산수량" not in cols:
                    continue
                rework_raw = ws.cell(row=r, column=cols["리웍"]).value
                if not isinstance(rework_raw, str) or not rework_raw.strip():
                    continue

                person = to_number(ws.cell(row=r, column=cols["인원"]).value)
                prod_qty = to_number(ws.cell(row=r, column=cols["생산수량"]).value)
                if person is None or person <= 0 or prod_qty is None:
                    continue

                plan_qty = to_number(ws.cell(row=r, column=cols["계획수량"]).value) if "계획수량" in cols else None
                defect_qty = to_number(ws.cell(row=r, column=cols["불량수량"]).value) if "불량수량" in cols else None

                category = classify_rework(rework_raw)
                ratio = prod_qty / person
                # 계획수량이 숫자로 안 읽힌 경우(빈칸, '-' 텍스트 등)는 plan_ratio를
                # None으로 둔다. 실적 기준 분석(ratio)에는 영향 없고, 계획 기준
                # 분석(plan_ratio)을 집계할 때만 이런 행을 걸러내면 된다.
                plan_ratio = plan_qty / person if plan_qty is not None else None

                records.append(
                    Record(
                        source_file=fname,
                        sheet=sheet_name,
                        month=month,
                        day=day,
                        product_group=str(product_group) if product_group is not None else "",
                        product_code=str(product_code) if product_code is not None else "",
                        product_name=str(product_name) if product_name is not None else "",
                        plan_qty=plan_qty,
                        prod_qty=prod_qty,
                        defect_qty=defect_qty,
                        rework_raw=rework_raw.strip(),
                        category=category,
                        person=person,
                        ratio=ratio,
                        plan_ratio=plan_ratio,
                    )
                )
    return records


def load_all_records(input_dir: str) -> pd.DataFrame:
    """input_dir 폴더의 모든 .XLSX 파일을 찾아 extract_records()로 읽고,
    하나의 pandas DataFrame으로 합친 뒤 중복을 제거해서 반환한다.

    파일 탐색:
      - Windows는 대소문자를 구분하지 않는 파일시스템이라 '*.XLSX'와
        '*.xlsx' glob이 같은 파일을 두 번 찾아올 수 있다. os.normcase(경로)를
        딕셔너리 키로 써서 같은 실제 경로는 한 번만 남긴다.
      - 엑셀이 열려 있으면 생기는 잠금 임시파일('~$파일명.XLSX')은 진짜
        데이터가 아니므로 제외한다.

    파일별 오류 처리:
      - 한 파일 처리 중 예외가 나도 전체가 멈추지 않도록 try/except로 감싸고,
        경고만 출력한 뒤 나머지 파일 처리를 계속한다.

    중복 제거:
      - 이 폴더의 파일들은 전부 '같은 마스터 생산계획표'를 날짜별로 저장해둔
        스냅샷이다. 즉 예를 들어 6/23 파일과 7/6 파일 양쪽에 6월 3일치
        데이터가 완전히 동일하게 들어있는 경우가 매우 많다(실제로 처음 돌려
        보니 원본 3833건 중 3458건이 이런 중복이었음). 시트/품번/품명/월/일/
        계획수량/생산수량/리웍텍스트/인원이 모두 같은 행은 같은 관측치로
        보고 하나만 남긴다. (반대로 같은 날짜인데 값이 수정된 경우는 서로
        다른 값으로 남기 때문에 dedup 키에 안 걸리고 둘 다 남는다 - 이는
        의도된 동작이며, 최신 스냅샷이 항상 우선하도록 만들지는 않았다.)
    """
    files = sorted(
        {
            os.path.normcase(f): f
            for f in glob.glob(os.path.join(input_dir, "*.XLSX")) + glob.glob(os.path.join(input_dir, "*.xlsx"))
            if not os.path.basename(f).startswith("~$")
        }.values()
    )
    if not files:
        raise FileNotFoundError(f"엑셀 파일을 찾지 못했습니다: {input_dir}")

    print(f"[정보] 엑셀 파일 {len(files)}개 발견")

    all_records: list[Record] = []
    for fp in files:
        try:
            recs = extract_records(fp)
            print(f"  - {os.path.basename(fp)}: {len(recs)}건 추출")
            all_records.extend(recs)
        except Exception as e:
            print(f"  ! {os.path.basename(fp)} 처리 중 오류 (건너뜀): {e}")

    df = pd.DataFrame([r.__dict__ for r in all_records])
    if df.empty:
        raise RuntimeError("추출된 데이터가 없습니다. 엑셀 구조가 예상과 다를 수 있습니다.")

    # 여러 파일이 같은 마스터 표의 날짜별 스냅샷이라 동일한 항목이 중복 등장한다.
    # 완전히 같은 내용(시트/품번/월/일/계획/생산/리웍/인원)이면 한 건으로 취급한다.
    dedup_keys = [
        "sheet", "product_code", "product_name", "month", "day",
        "plan_qty", "prod_qty", "rework_raw", "person",
    ]
    before = len(df)
    df = df.drop_duplicates(subset=dedup_keys).reset_index(drop=True)
    print(f"[정보] 중복 제거: {before}건 -> {len(df)}건")

    return df


def summarize(df: pd.DataFrame, ratio_col: str = "ratio", qty_col: str = "prod_qty") -> pd.DataFrame:
    """카테고리(리웍 분류)별로 1인당 수량 비율의 통계 요약표를 만든다.

    두 가지 기준으로 재사용한다:
      - 실적 기준: ratio_col="ratio" (=생산수량/인원), qty_col="prod_qty"
      - 계획 기준: ratio_col="plan_ratio" (=계획수량/인원), qty_col="plan_qty"
    ratio_col 값이 None/NaN인 행(계획수량이 숫자로 안 읽힌 경우 등)은
    dropna로 미리 제외하고 집계한다.

    건수/평균/중앙값/표준편차/최소/최대에 더해, 총수량과 총인원(단순 합계)도
    같이 붙이고, 마지막으로 변동계수(표준편차/평균)를 계산한다. 변동계수가
    클수록 그 공정은 날짜/건마다 1인당 수량 편차가 크다는 뜻이라, 원래
    질문("어떤 날은 인원이 너무 많이 투입되는 것 아니냐")에 대한 핵심
    지표다. CATEGORY_ORDER로 카테고리를 정렬된 Categorical로 바꿔서, 요약표와
    그래프에서 항상 같은 순서(셀라인→단발→로타리→SET→접지→충진→포장→기타)로
    나오게 한다.
    """
    df = df.dropna(subset=[ratio_col]).copy()
    df["category"] = pd.Categorical(df["category"], categories=CATEGORY_ORDER, ordered=True)
    g = df.groupby("category", observed=True)[ratio_col]
    summary = g.agg(
        건수="count",
        평균=("mean"),
        중앙값=("median"),
        표준편차=("std"),
        최소=("min"),
        최대=("max"),
    )
    totals = df.groupby("category", observed=True).agg(
        총수량=(qty_col, "sum"), 총인원=("person", "sum")
    )
    summary = summary.join(totals)
    summary["표준편차/평균(변동계수)"] = summary["표준편차"] / summary["평균"]
    return summary.round(2)


def summarize_by_group(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """제품군(마스크/튜브/용기 등) x 리웍 카테고리 교차표를 두 가지 형태로 만든다.

    - counts: (제품군, 카테고리)별 원본 건수 (pd.crosstab)
    - pct: 같은 표를 '행(제품군) 기준 비율(%)'로 정규화한 것. 즉 각 행의
      합이 100%가 되어, "이 제품군에서 작업한 건 중 몇 %가 어느 라인/공정
      이었는가"를 바로 볼 수 있다. (제품군마다 전체 건수가 다르므로 원본
      건수만 보면 제품군 간 비교가 어려워서 비율표를 따로 만든다.)
    product_group의 순서는 전체 건수가 많은 순(value_counts 기준)으로,
    category의 순서는 CATEGORY_ORDER로 고정해서 요약표/히트맵이 항상 같은
    순서로 보이게 한다.
    """
    df = df.copy()
    df["category"] = pd.Categorical(df["category"], categories=CATEGORY_ORDER, ordered=True)
    group_order = [g for g in df["product_group"].value_counts().index]
    df["product_group"] = pd.Categorical(df["product_group"], categories=group_order, ordered=True)

    counts = pd.crosstab(df["product_group"], df["category"], dropna=False)
    counts = counts.reindex(columns=CATEGORY_ORDER, fill_value=0)
    pct = counts.div(counts.sum(axis=1), axis=0).mul(100).round(1)
    return counts, pct


def print_etc_breakdown(df: pd.DataFrame, top_n: int = 20):
    """'기타'로 분류된 리웍 원본 텍스트들을 빈도순으로 콘솔에 출력한다.

    classify_rework()의 키워드에 안 걸린(두라인/잔업/재고메모 등) 텍스트가
    무엇이고 얼마나 자주 나오는지 보여줘서, 나중에 특정 표기가 꽤 잦다면
    CATEGORY_KEYWORDS에 새 카테고리로 추가할지 판단할 수 있게 하기 위한
    참고용 출력이다. (분석 결과 자체에는 영향 없음, 그냥 확인용.)
    """
    etc = df[df["category"] == "기타"]
    if etc.empty:
        return
    counts = Counter(etc["rework_raw"])
    print("\n[참고] '기타'로 분류된 리웍 텍스트 상위 목록 (필요시 CATEGORY_KEYWORDS에 추가 고려):")
    for text, cnt in counts.most_common(top_n):
        print(f"    {text!r}: {cnt}건")


def _ratio_boxplot_and_bar(
    df: pd.DataFrame,
    ratio_col: str,
    ylabel: str,
    title_metric: str,
    output_dir: str,
    box_filename: str,
    bar_filename: str,
    colors,
) -> tuple[str, str]:
    """카테고리별 1인당 수량 비율(ratio_col)의 박스플롯 + 평균 막대그래프
    한 쌍을 그려서 파일로 저장하고, 저장된 두 파일 경로를 반환한다.

    make_plots()에서 '생산수량 기준'과 '계획수량 기준' 두 번 호출해서 쓰는
    공통 로직이다. ratio_col이 None/NaN인 행은 미리 제외한 뒤 카테고리별로
    묶는다.
      - 박스플롯: 중앙값/사분위수/이상치를 보여주고, 그 위에 실제 관측치들을
        약간 흔들어(jitter) 점으로 겹쳐서 "박스 뒤에 숨은 실제 표본 개수와
        분포 모양"까지 같이 보이게 한다.
      - 막대그래프: 평균 + 표준편차(오차막대)만 뽑아 한눈에 비교할 수 있게
        요약한 보조 그래프.
    """
    sub = df.dropna(subset=[ratio_col])
    categories = [c for c in CATEGORY_ORDER if c in sub["category"].unique()]

    fig, ax = plt.subplots(figsize=(10, 6))
    data = [sub.loc[sub["category"] == c, ratio_col].values for c in categories]
    bp = ax.boxplot(data, tick_labels=categories, showmeans=True, patch_artist=True)
    for patch, color in zip(bp["boxes"], colors):
        patch.set_facecolor(color)
        patch.set_alpha(0.35)
    rng = np.random.default_rng(0)
    for i, c in enumerate(categories):
        y = sub.loc[sub["category"] == c, ratio_col].values
        x = rng.normal(i + 1, 0.06, size=len(y))
        ax.scatter(x, y, s=14, alpha=0.5, color=colors[i % len(colors)])
    ax.set_ylabel(ylabel)
    ax.set_title(f"공정(리웍)별 1인당 {title_metric} 분포 - 날짜 무시, 전체 건 기준")
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    fig.tight_layout()
    p_box = os.path.join(output_dir, box_filename)
    fig.savefig(p_box, dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(9, 6))
    means = [sub.loc[sub["category"] == c, ratio_col].mean() for c in categories]
    stds = [sub.loc[sub["category"] == c, ratio_col].std() for c in categories]
    ax.bar(categories, means, yerr=stds, capsize=6, color=colors[: len(categories)], alpha=0.8)
    ax.set_ylabel(f"평균 1인당 {title_metric}")
    ax.set_title(f"공정(리웍)별 평균 1인당 {title_metric} (오차막대: 표준편차)")
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    fig.tight_layout()
    p_bar = os.path.join(output_dir, bar_filename)
    fig.savefig(p_bar, dpi=150)
    plt.close(fig)

    return p_box, p_bar


def make_plots(df: pd.DataFrame, output_dir: str):
    """분석 결과를 그래프 7장으로 그려 output_dir에 PNG로 저장한다.

    df는 이미 record 단위(제품 x 하루)로 정리되고 category가 매겨진
    DataFrame이다. 그래프는 전부 '날짜를 구분하지 않고' 카테고리(라인/공정)
    기준으로만 묶어서, 같은 공정 안에서도 건마다 1인당 수량이 얼마나
    들쭉날쭉한지가 보이도록 그린다. 01~02는 실제 생산수량 기준, 06~07은
    계획수량 기준으로 같은 형태의 그래프를 한 번 더 그려서 "계획 대비
    실적이 유독 낮았던 건지, 애초에 계획부터 헐거웠던 건지"를 비교할 수
    있게 한다.
    """
    os.makedirs(output_dir, exist_ok=True)
    # CATEGORY_ORDER 순서를 유지하면서, 실제로 데이터가 있는 카테고리만 남긴다.
    categories = [c for c in CATEGORY_ORDER if c in df["category"].unique()]
    colors = plt.get_cmap("tab10").colors

    # 1~2) 실적(생산수량) 기준 1인당 수량 박스플롯 + 평균 막대그래프.
    #    -> 박스가 넓거나 점이 넓게 퍼진 카테고리일수록 "그날그날 인원 대비
    #       생산량 편차가 크다"는 뜻이라, 원래 목적(인원 과다투입 날 찾기)에
    #       가장 직접적으로 답하는 핵심 그래프.
    p1, p2 = _ratio_boxplot_and_bar(
        df, "ratio", "생산수량 / 인원 (1인당 생산수량)", "생산수량",
        output_dir, "01_category_ratio_boxplot.png", "02_category_ratio_mean_bar.png", colors,
    )

    # 3) 인원수 vs 생산수량 산점도 (카테고리별 색상 구분).
    #    ratio(기울기)로 뭉개지 않고 원본 두 축(인원, 생산량)을 그대로 보여줘서,
    #    "인원은 비슷한데 생산량만 유독 낮은 점"처럼 raw 데이터 상에서 이상치를
    #    직접 눈으로 짚어낼 수 있게 하는 보조 그래프.
    fig, ax = plt.subplots(figsize=(9, 7))
    for i, c in enumerate(categories):
        sub = df[df["category"] == c]
        ax.scatter(sub["person"], sub["prod_qty"], s=18, alpha=0.5, label=c, color=colors[i % len(colors)])
    ax.set_xlabel("투입 인원")
    ax.set_ylabel("생산수량")
    ax.set_title("투입 인원 대비 생산수량 (공정별)")
    ax.legend()
    ax.grid(linestyle="--", alpha=0.4)
    fig.tight_layout()
    p3 = os.path.join(output_dir, "03_person_vs_production_scatter.png")
    fig.savefig(p3, dpi=150)
    plt.close(fig)

    # 4) 카테고리별 데이터 건수 막대그래프.
    #    1~3번 그래프를 해석할 때 "이 카테고리는 표본이 몇 개 안 돼서 통계가
    #    불안정할 수 있다"를 함께 판단하기 위한 참고 그래프. 예: '단발'/'접지'
    #    처럼 건수가 적은 카테고리는 평균/표준편차가 소수의 극단값에 쉽게
    #    흔들릴 수 있음.
    fig, ax = plt.subplots(figsize=(8, 5))
    counts = [ (df["category"] == c).sum() for c in categories]
    ax.bar(categories, counts, color=colors[: len(categories)], alpha=0.8)
    ax.set_ylabel("건수")
    ax.set_title("공정(리웍)별 데이터 건수")
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    fig.tight_layout()
    p4 = os.path.join(output_dir, "04_category_count_bar.png")
    fig.savefig(p4, dpi=150)
    plt.close(fig)

    # 5) 제품군(마스크/튜브/용기 등) x 리웍(카테고리) 히트맵.
    #    summarize_by_group()이 계산한 '행(제품군) 기준 비율(%)' 표를 색으로
    #    표현하고, 각 칸에는 비율(%)과 원본 건수를 함께 적는다. 색이 진할수록
    #    그 제품군에서 해당 라인/공정이 차지하는 비중이 크다는 뜻이라, 한눈에
    #    "어느 제품군은 어느 라인을 주로 쓰는지"를 파악할 수 있다.
    counts, pct = summarize_by_group(df)
    groups = list(counts.index)
    cats = [c for c in CATEGORY_ORDER if c in counts.columns]
    counts = counts[cats]
    pct = pct[cats]

    fig, ax = plt.subplots(figsize=(1.4 * len(cats) + 2, 1.1 * len(groups) + 2))
    im = ax.imshow(pct.values, cmap="YlOrRd", aspect="auto", vmin=0)
    ax.set_xticks(range(len(cats)))
    ax.set_xticklabels(cats, rotation=30, ha="right")
    ax.set_yticks(range(len(groups)))
    ax.set_yticklabels(groups)
    for i in range(len(groups)):
        for j in range(len(cats)):
            p = pct.values[i, j]
            n = counts.values[i, j]
            # 배경이 진한 색(비율이 높은 칸)일 때는 흰 글씨로 바꿔서 가독성을 유지한다.
            text_color = "white" if p > pct.values.max() * 0.6 else "black"
            ax.text(j, i, f"{p:.0f}%\n({n}건)", ha="center", va="center", fontsize=9, color=text_color)
    ax.set_title("제품군별 리웍(공정) 사용 비중 - 행(제품군) 기준 %")
    fig.colorbar(im, ax=ax, label="해당 제품군 내 비중(%)")
    fig.tight_layout()
    p5 = os.path.join(output_dir, "05_group_category_heatmap.png")
    fig.savefig(p5, dpi=150)
    plt.close(fig)

    # 6~7) 계획(계획수량) 기준 1인당 수량 박스플롯 + 평균 막대그래프.
    #    1~2번과 완전히 같은 형태지만 분자를 생산수량 대신 계획수량으로 바꿔서
    #    그린다. 계획수량이 비어있는 행은 _ratio_boxplot_and_bar 내부에서
    #    자동으로 제외된다. 1~2번(실적)과 나란히 비교하면 "인원 대비 계획량
    #    자체가 적었던 건지" 아니면 "계획은 정상인데 실적만 저조했는지"를
    #    구분해서 볼 수 있다.
    p6, p7 = _ratio_boxplot_and_bar(
        df, "plan_ratio", "계획수량 / 인원 (1인당 계획수량)", "계획수량",
        output_dir, "06_category_planratio_boxplot.png", "07_category_planratio_mean_bar.png", colors,
    )

    print(f"\n[정보] 그래프 저장 완료: {output_dir}")
    for p in [p1, p2, p3, p4, p5, p6, p7]:
        print(f"  - {p}")


def main():
    """전체 파이프라인을 순서대로 실행하는 진입점.

    흐름: 엑셀 폴더 읽기(load_all_records) -> 원본 데이터 CSV 저장 ->
    카테고리별 요약을 실적(생산수량) 기준과 계획(계획수량) 기준 두 번
    계산/저장/출력(summarize) -> '기타' 분류 상세 참고 출력 -> 제품군x
    카테고리 교차표(summarize_by_group) 계산/저장/출력 -> 그래프 7장 생성
    (make_plots). 각 단계 결과는 output-dir(기본값: 이 스크립트가 있는
    폴더의 'output' 하위 폴더)에 CSV/PNG로 남겨서, 콘솔을 스크롤하지 않아도
    나중에 다시 열어볼 수 있게 한다.
    """
    parser = argparse.ArgumentParser(description="일일생산계획표 리웍(공정)별 1인당 생산수량 분석")
    parser.add_argument("--input-dir", default=DEFAULT_INPUT_DIR, help="엑셀 파일들이 있는 폴더")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="결과(csv/그래프) 저장 폴더")
    args = parser.parse_args()

    # 1) 모든 엑셀 파일 -> 레코드 단위 DataFrame (중복 제거까지 완료된 상태)
    df = load_all_records(args.input_dir)

    # 2) 원본 레코드를 그대로 CSV로 남긴다 (엑셀에서 직접 검증하고 싶을 때 사용).
    os.makedirs(args.output_dir, exist_ok=True)
    raw_csv = os.path.join(args.output_dir, "records.csv")
    df.to_csv(raw_csv, index=False, encoding="utf-8-sig")
    print(f"[정보] 원본 추출 데이터 저장: {raw_csv} ({len(df)}행)")

    # 3) 카테고리(리웍 분류)별 1인당 생산수량(실적 기준) 통계 요약.
    summary = summarize(df, ratio_col="ratio", qty_col="prod_qty")
    summary_csv = os.path.join(args.output_dir, "summary_by_category.csv")
    summary.to_csv(summary_csv, encoding="utf-8-sig")
    print(f"\n[결과] 공정(리웍)별 1인당 생산수량(실적 기준) 요약:\n{summary}\n")
    print(f"[정보] 요약 데이터 저장: {summary_csv}")

    # 3-1) 같은 집계를 계획수량 기준(plan_ratio = 계획수량/인원)으로도 계산.
    #    실적 기준과 나란히 비교하면 "계획 단계부터 인원 대비 계획량이
    #    적었던 공정"과 "계획은 괜찮았는데 실적만 저조했던 공정"을 구분할
    #    수 있다.
    plan_summary = summarize(df, ratio_col="plan_ratio", qty_col="plan_qty")
    plan_summary_csv = os.path.join(args.output_dir, "summary_by_category_plan.csv")
    plan_summary.to_csv(plan_summary_csv, encoding="utf-8-sig")
    print(f"[결과] 공정(리웍)별 1인당 계획수량(계획 기준) 요약:\n{plan_summary}\n")
    print(f"[정보] 요약 데이터 저장: {plan_summary_csv}")

    # 4) '기타'로 빠진 텍스트들이 뭔지 참고용으로 출력 (카테고리 키워드 보강 판단용).
    print_etc_breakdown(df)

    # 5) 제품군(마스크/튜브/용기) x 리웍 카테고리 교차표: 건수표 + 비율표.
    group_counts, group_pct = summarize_by_group(df)
    group_counts_csv = os.path.join(args.output_dir, "group_category_counts.csv")
    group_pct_csv = os.path.join(args.output_dir, "group_category_pct.csv")
    group_counts.to_csv(group_counts_csv, encoding="utf-8-sig")
    group_pct.to_csv(group_pct_csv, encoding="utf-8-sig")
    print(f"\n[결과] 제품군별 리웍 건수:\n{group_counts}\n")
    print(f"[결과] 제품군별 리웍 비중(%):\n{group_pct}\n")
    print(f"[정보] 제품군x리웍 집계 저장: {group_counts_csv}, {group_pct_csv}")

    # 6) 위 데이터를 바탕으로 그래프 7장 생성 (실적 기준 2 + 산점도 1 + 건수 1
    #    + 히트맵 1 + 계획 기준 2).
    make_plots(df, args.output_dir)


if __name__ == "__main__":
    main()

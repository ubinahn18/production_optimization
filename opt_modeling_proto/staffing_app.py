# -*- coding: utf-8 -*-
"""
staffing_app.py

30일 최적화 결과(line_schedule.csv)를 바탕으로, 특정 하루 안에서 "실제
인원 배치"를 수작업으로 해볼 수 있는 로컬 웹 도구. gui_app.py(전체 30일
계획을 CP-SAT으로 돌리는 도구)와는 별개로, 이미 나온 계획을 놓고 사람이
직접 라인 배정/인원 편성을 조정하는 용도다 - 여기엔 최적화나 알고리즘이
전혀 없다. 전부 사람이 드래그앤드롭으로 직접 조정하고, 그 결과를 저장/
확인하는 것뿐이다.

핵심 개념 두 가지:
  1) 슬롯 교환: 같은 슬롯(시간) 안에서는 어느 라인이 뭘 하든 그 슬롯의
     총 필요인원이 안 바뀌므로 자유롭게 맞바꿀 수 있다. 다른 슬롯끼리는
     "필요인원이 똑같은 셀끼리만" 맞바꿀 수 있다(그래야 그 두 슬롯의
     총 필요인원이 각각 그대로 유지됨). 셋업 셀은 필요인원이 항상
     0이므로 사실상 아무 0인원 셀과나 자유롭게 바뀌고, 그냥 idle로
     지우는 것도 가능하게 해준다(검증 없이 사람이 알아서 판단).
  2) 인원 "블록": 그날 총 고용인원을 사람이 원하는 크기로 나눠서 여러
     팀(블록)을 만들고, 각 팀을 슬롯별로 어느 라인에 배치할지 드래그로
     정한다. 한 슬롯의 필요인원을 여러 블록을 합쳐서 채울 수 있다(예:
     6명 블록 + 5명 블록 = 11명 슬롯). 배치가 끝나면 블록별로 하루 동안
     어느 라인을 돌았는지 타임라인으로 볼 수 있다.

사용법:
    python staffing_app.py --real-plan --read-specs excel --reference-date 2026-08-03
    (그 뒤 브라우저에서 http://127.0.0.1:5000 접속)

--real-plan 등 소스 관련 인자는 output/_data_source.py의 add_source_args와
동일하다(plot_day.py/labor_utilization.py와 같은 방식) - line_schedule.csv/
daily_workforce.csv는 그대로 읽고, 라인별 필요인원(workers)만 다시 계산
(엑셀 재로딩 필요, 시간이 좀 걸림 - 서버 기동 시 한 번만).

기본적으로 HTTP Basic Auth 비밀번호 보호가 켜져 있다(기본 비밀번호는
_DEFAULT_PASSWORD 참고, --password로 바꾸거나 --no-password로 끌 수
있음) - ngrok 등으로 외부에 링크를 공유해도 비밀번호 없이는 못 들어옴.
"""

from __future__ import annotations

import argparse
import colorsys
import hashlib
import io
import json
import os
import re
import secrets
import sys

import pandas as pd
from flask import Flask, jsonify, request, send_file, send_from_directory

# opt_modeling_proto/ 자체(현재 파일 위치)와 output/(스크립트 폴더, 패키지
# 아님)를 둘 다 import 경로에 추가해야 _data_source.py와 scheduling 패키지를
# 가져올 수 있다.
_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _HERE)
sys.path.insert(0, os.path.join(_HERE, "output"))

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace", line_buffering=True)

from _data_source import add_source_args, build_workers_lookup, resolve_source  # noqa: E402
from scheduling.models import SLOT_LABELS  # noqa: E402
from scheduling.report import _natural_line_key  # noqa: E402

app = Flask(__name__, static_folder=os.path.join(_HERE, "staffing_static"), static_url_path="")

# 사람이 원하는 라인 표시 순서(카테고리 단위) - 이 목록에 없는 카테고리는
# 뒤로 밀린다. 같은 카테고리 안에서는 _natural_line_key로 번호순 정렬
# (예: 단발_2가 단발_10보다 앞).
_LINE_CATEGORY_ORDER = ["셀라인", "단발", "10열기", "로타리", "튜브라인"]


def _line_category(lid: str) -> str:
    """line_id에서 끝의 '_숫자'를 뗀 카테고리 이름(예: '단발_10' -> '단발').
    끝에 '_숫자'가 없는 라인(예: '대용량파우치')은 그 자체가 카테고리."""
    m = re.match(r"^(.*?)_(\d+)$", lid)
    return m.group(1) if m else lid


def _line_sort_key(lid: str):
    category = _line_category(lid)
    try:
        rank = _LINE_CATEGORY_ORDER.index(category)
    except ValueError:
        rank = len(_LINE_CATEGORY_ORDER)
    return (rank, _natural_line_key(lid))

# --password로 지정했을 때만 켜지는 아주 단순한 접근 보호(HTTP Basic
# Auth). ngrok 등으로 외부에 링크를 공유할 때, 그 링크를 우연히/실수로
# 본 사람이 바로 데이터를 보거나 고칠 수 없게 최소한의 문턱을 둔다.
# 아이디는 아무거나 상관없고 비밀번호만 확인한다.
_PASSWORD: str | None = None

# 외부 공유(ngrok 등) 시 매번 --password를 타이핑하지 않아도 되도록 둔
# 기본 비밀번호. --no-password로 끌 수 있고, --password로 다른 값을 줄
# 수도 있다.
_DEFAULT_PASSWORD = "cCQnf4zvTRFAsK"


@app.before_request
def _check_password():
    if not _PASSWORD:
        return None
    auth = request.authorization
    if not auth or not secrets.compare_digest(auth.password or "", _PASSWORD):
        return (
            "인증이 필요합니다.",
            401,
            {"WWW-Authenticate": 'Basic realm="staffing"'},
        )
    return None

# 모듈 전역 상태 - 이 도구는 개인 로컬 도구라 단일 사용자/단일 프로세스
# 가정하고 단순하게 전역 변수에 들고 있는다(멀티유저 동시접속 고려 안 함).
STATE: dict = {
    "schedule_df": None,       # line_schedule.csv 전체
    "workforce_df": None,      # daily_workforce.csv 전체
    "workers_lookup": None,    # {(line_id, order_id): workers}
    "state_dir": None,         # 편집 상태 저장용 디렉터리(line_schedule.csv와 같은 폴더 밑 staffing_state/)
    "all_lines": None,         # 자연정렬된 전체 물리 라인 목록(36개 등) - 정렬 기준 및
                               # "그날 활동 있는 라인만" 필터링의 기준 순서로 쓰임(day
                               # 자체의 표시 목록은 아님, _build_fresh_day_state 참고)
    "schedule_fingerprint": None,  # line_schedule.csv 내용의 sha256 - 저장된 날짜별 상태가
                                    # "지금 로드된 계획"과 같은 계획을 기준으로 만들어진 건지
                                    # 판별하는 데 씀(아래 api_get_day 참고). plan_from_orders.py를
                                    # 다시 돌려서 계획이 바뀌면 이 값도 바뀌므로, 예전에 저장해둔
                                    # 날짜를 열어도 그 예전 스케줄이 아니라 최신 스케줄 기준으로
                                    # 새로 만들어진다(사람이 하루하루 초기화 안 눌러도 됨) -
                                    # 2026-07-21: "새 스케줄 돌리면 인원배치 도구도 그 스케줄
                                    # 기준으로 맞춰져야 한다"는 실제 요청 반영.
}


def _state_path(day: int) -> str:
    return os.path.join(STATE["state_dir"], f"day{day}.json")


def _build_fresh_day_state(day: int) -> dict:
    """line_schedule.csv/daily_workforce.csv/workers_lookup에서 이 날짜의
    초기 상태(사람이 아직 아무것도 안 건드린 상태)를 만든다. 블록/배치는
    빈 채로 시작한다."""
    df = STATE["schedule_df"]
    day_df = df[df["day"] == day]
    if day_df.empty:
        raise ValueError(f"day={day}에 해당하는 행이 없습니다.")

    by_line_slot: dict[tuple[str, str], dict] = {}
    active_lines: set[str] = set()
    for row in day_df.itertuples(index=False):
        by_line_slot[(row.line_id, row.slot)] = {
            "activity": row.activity,
            "order_id": "" if pd.isna(row.order_id) else str(row.order_id),
            "product_id": "" if pd.isna(row.product_id) else str(row.product_id),
        }
        if row.activity != "idle":
            active_lines.add(row.line_id)

    # 그날 하루 종일 대기(idle)만 한 라인은 화면에서 뺀다 - plot_day.py와
    # 같은 필터. 물리 라인이 36대라 다 보여주면 화면 세로 길이가 너무
    # 길어져서 한눈에 안 들어오고, 어차피 그날 아무 일도 안 한 라인으로
    # 옮길 일은 이 도구의 용도(그날 이미 배정된 작업의 재배치)상 거의 없다.
    lines_for_day = [lid for lid in STATE["all_lines"] if lid in active_lines]

    workers_lookup = STATE["workers_lookup"]
    grid: dict[str, list[dict]] = {}
    for line_id in lines_for_day:
        cells = []
        for slot in SLOT_LABELS:
            cell = by_line_slot.get((line_id, slot), {"activity": "idle", "order_id": "", "product_id": ""})
            workers = 0
            if cell["activity"] == "produce" and cell["order_id"]:
                workers = workers_lookup.get((line_id, cell["order_id"]))
                if workers is None:
                    workers = 0
            cells.append({
                "activity": cell["activity"],
                "order_id": cell["order_id"],
                "product_id": cell["product_id"],
                "workers": int(workers),
            })
        grid[line_id] = cells

    wf_row = STATE["workforce_df"]
    wf_row = wf_row[wf_row["day"] == day]
    daily_headcount = int(wf_row.iloc[0]["workforce"]) if not wf_row.empty else 0

    return {
        "day": day,
        "slot_labels": SLOT_LABELS,
        "daily_headcount": daily_headcount,
        "lines": lines_for_day,
        "grid": grid,
        # flows: {"t": [{"id":.., "srcLine":.., "dstLine":.., "count":..}, ...]}
        # 슬롯 t(0-idx)에서 t+1로 넘어갈 때 어느 라인(또는 "__idle__" =
        # 미배치/휴식)에서 어느 라인으로 몇 명이 이동하는지. 슬롯이
        # len(SLOT_LABELS)개면 전환은 그보다 하나 적게 존재한다.
        "flows": {str(t): [] for t in range(len(SLOT_LABELS) - 1)},
        # closed_lines: 사람이 "이건 하루 종일 뻔한 연속생산이라 슬롯
        # 전환 편집에서 안 봐도 된다"고 접어둔 라인 목록 - 그림에서도
        # 행이 숨겨진다(엑셀 행숨김과 동일한 개념, 체크 해제하면 복귀).
        "closed_lines": [],
        # tracking: {tree_id: {id, label, color, rootNodeId, nodes: {node_id: {
        #   count, lineId, slotIdx, edgeId, parentId, childIds}}}} - 특정
        # 인원 부분집합이 하루 동안 화살표를 따라 어떻게 갈라져 이동했는지
        # 사람이 직접 클릭해서 추적해둔 트리(들). 실제 배정과는 무관한
        # 순수 주석/추적용 데이터.
        "tracking": {},
        # 이 상태가 어느 line_schedule.csv 내용을 기준으로 만들어졌는지
        # (api_get_day의 최신성 검사에서 씀).
        "schedule_fingerprint": STATE["schedule_fingerprint"],
    }


def _migrate_tracking_add_start_nodes(data: dict) -> None:
    """예전 버전에서 만들어진 추적 트리는 뿌리 노드가 "첫 화살표의
    도착점"이었다(그 화살표의 출발 슬롯 자체는 트리에 없었음). 뿌리의
    출발 라인을 아래 순서로 찾아서, 찾아지면 뿌리 앞에 그 출발점을
    실제 노드로 끼워넣어 새 스키마와 똑같이 맞춘다:
      1) 뿌리에 srcLine 필드가 이미 있으면 그걸 씀(중간 버전 데이터).
      2) 없으면(가장 옛날 데이터), 뿌리의 edgeId가 가리키는 화살표가
         아직 flows에 남아있는지 찾아서 그 화살표의 srcLine을 씀.
      3) 그것도 못 찾으면(화살표가 이미 삭제됨 등) 알 방법이 없으므로
         그대로 둔다.
    이미 새 스키마인 트리(뿌리에 edgeId 자체가 없음)는 건드리지 않고,
    한 번 마이그레이션된 트리는 재실행해도 다시 안 건드린다.
    """
    flows = data.get("flows") or {}
    for tree in (data.get("tracking") or {}).values():
        nodes = tree.get("nodes") or {}
        root_id = tree.get("rootNodeId")
        root = nodes.get(root_id)
        if not root or not root.get("edgeId"):
            continue  # 이미 새 스키마(시작 노드가 뿌리, edgeId 없음)

        src_line = root.get("srcLine")
        if src_line is None:
            transition_key = str(root["slotIdx"] - 1)
            for e in flows.get(transition_key, []):
                if e.get("id") == root["edgeId"]:
                    src_line = e.get("srcLine")
                    break
        if src_line is None:
            continue  # 정말로 알 방법이 없음(화살표가 이미 삭제됐거나 수정됨)

        start_id = f"migrated_start_{root_id}"
        if start_id in nodes:
            continue
        nodes[start_id] = {
            "id": start_id, "count": root.get("count", 0),
            "lineId": src_line, "slotIdx": root["slotIdx"] - 1,
            "edgeId": None, "parentId": None, "childIds": [root_id],
        }
        root["parentId"] = start_id
        tree["rootNodeId"] = start_id


def _migrate_saved_state(data: dict) -> dict:
    """예전 저장 형식을 지금 스키마로 보정한다(둘 다 없으면 그대로,
    있으면 그 부분만 채워넣음 - 매번 전부 다시 만들 필요 없음).

    1) 블록 방식(고정 크기 "블록"을 슬롯별로 라인에 드래그해 배치하던
       예전 UI) -> flows(슬롯 전환별 이동 배정)로 변환. 블록 하나가
       슬롯 t에 라인 X, 슬롯 t+1에 라인 Y에 있었다면, "X에서 Y로 그
       블록 인원만큼 이동"이라는 흐름 한 개로 그대로 옮겨진다(X==Y면
       "그 자리에 계속 있었다"는 self-loop 흐름, None이면 미배치/휴식).
    2) closed_lines(접어둔 라인 체크박스) 필드가 없으면(그 기능이
       생기기 전에 저장된 파일) 빈 목록으로 채워넣음.
    3) tracking(인원 추적 트리) 필드가 없으면 빈 dict로 채워넣음.
    """
    if "flows" not in data:
        blocks = data.get("blocks", [])
        assignments = data.get("assignments", {})
        slot_labels = data.get("slot_labels", SLOT_LABELS)
        n = len(slot_labels)
        flows: dict[str, list[dict]] = {str(t): [] for t in range(n - 1)}
        for b in blocks:
            arr = assignments.get(b["id"], [None] * n)
            for t in range(n - 1):
                src = arr[t] if t < len(arr) else None
                dst = arr[t + 1] if t + 1 < len(arr) else None
                flows[str(t)].append({
                    "id": f"migrated_{b['id']}_{t}",
                    "srcLine": src or "__idle__",
                    "dstLine": dst or "__idle__",
                    "count": b["size"],
                })
        data["flows"] = flows
        data.pop("blocks", None)
        data.pop("assignments", None)

    if "closed_lines" not in data:
        data["closed_lines"] = []

    if "tracking" not in data:
        data["tracking"] = {}

    _migrate_tracking_add_start_nodes(data)

    # "lines"는 화면에 보여줄 순서일 뿐 실제 배정 데이터(grid/flows)와는
    # 무관하므로, 표시 순서 기준(_line_sort_key)이 바뀌면 예전에 저장된
    # 날짜도 매번 최신 순서로 다시 정렬해서 보여준다(한 번만 채워넣는
    # 다른 필드들과 달리, 이건 매 로드마다 다시 적용).
    if "lines" in data:
        data["lines"] = sorted(data["lines"], key=_line_sort_key)

    return data


def _hash_hue(text: str) -> int:
    """app.js의 hashHue()와 같은 계산 - 라인 id마다 고정된 hue를 뽑는다."""
    h = 0
    for ch in text:
        h = (h * 31 + ord(ch)) & 0xFFFFFFFF
    return h % 360


def _line_fill_hex(line_id: str) -> str:
    """app.js의 productColor()(hsl(hue, 65%, 83%))와 같은 파스텔 색을
    엑셀 셀 배경색(PatternFill)에 쓸 RGB 16진수로 변환."""
    hue = _hash_hue(line_id) / 360.0
    r, g, b = colorsys.hls_to_rgb(hue, 0.83, 0.65)  # colorsys는 (h, l, s) 순서
    return "{:02X}{:02X}{:02X}".format(int(r * 255), int(g * 255), int(b * 255))


def _reconstruct_leaf_timetables(data: dict) -> list[dict]:
    """state.tracking의 각 추적 트리에서, 자식이 없는 노드(가지 끝)마다
    그 부분집합이 하루 동안 슬롯별로 어느 라인에 있었는지를 뿌리까지
    거슬러 올라가며 재구성한다. 뿌리 노드는 "도착한 슬롯/라인"만 들고
    있으므로, 뿌리 바로 이전 슬롯(추적을 시작한 화살표의 출발점)은
    노드에 같이 저장해둔 srcLine으로 채운다. 추적이 닿지 않은 슬롯은
    결과 딕셔너리에 아예 키가 없다(모른다는 뜻 - 빈 칸으로 표시).
    """
    results: list[dict] = []
    for tree in (data.get("tracking") or {}).values():
        nodes = tree.get("nodes") or {}
        leaf_ids = [nid for nid, nd in nodes.items() if not nd.get("childIds")]
        for leaf_id in leaf_ids:
            chain = []
            cur = nodes.get(leaf_id)
            while cur is not None:
                chain.append(cur)
                parent_id = cur.get("parentId")
                cur = nodes.get(parent_id) if parent_id else None
            if not chain:
                continue
            chain.reverse()  # 이제 뿌리 -> 리프 순서

            slot_line: dict[int, str] = {}
            root = chain[0]
            if root.get("srcLine") is not None:
                slot_line[root["slotIdx"] - 1] = root["srcLine"]
            for nd in chain:
                slot_line[nd["slotIdx"]] = nd["lineId"]

            leaf_node = chain[-1]
            results.append({
                "tree_label": tree.get("label", ""),
                "count": leaf_node.get("count", 0),
                "slot_line": slot_line,
            })
    return results


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/api/days")
def api_days():
    days = sorted(int(d) for d in STATE["schedule_df"]["day"].unique())
    saved = set()
    if os.path.isdir(STATE["state_dir"]):
        for fn in os.listdir(STATE["state_dir"]):
            if fn.startswith("day") and fn.endswith(".json"):
                try:
                    saved.add(int(fn[3:-5]))
                except ValueError:
                    pass
    return jsonify({"days": days, "saved": sorted(saved)})


@app.route("/api/day/<int:day>")
def api_get_day(day: int):
    path = _state_path(day)
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        # 저장된 날짜가 "지금 로드된 계획"과 다른 계획(plan_from_orders.py를
        # 다시 돌려서 line_schedule.csv 내용 자체가 바뀐 경우) 기준으로
        # 만들어졌다면, 그 예전 저장을 그대로 보여주지 않고 최신 계획
        # 기준으로 새로 만든다 - 서버를 재시작만 한 경우(계획은 그대로)는
        # fingerprint가 안 바뀌므로 저장된 내용이 정상적으로 그대로
        # 유지된다. 예전 파일은 지우지 않는다 - 다시 저장하면 그때
        # 최신 계획 기준 내용으로 자연스럽게 덮어써진다.
        if data.get("schedule_fingerprint") != STATE["schedule_fingerprint"]:
            return jsonify(_build_fresh_day_state(day))
        data = _migrate_saved_state(data)
        return jsonify(data)
    return jsonify(_build_fresh_day_state(day))


@app.route("/api/day/<int:day>/state", methods=["POST"])
def api_save_day(day: int):
    data = request.get_json(force=True)
    os.makedirs(STATE["state_dir"], exist_ok=True)
    with open(_state_path(day), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    return jsonify({"ok": True})


@app.route("/api/day/<int:day>/reset", methods=["POST"])
def api_reset_day(day: int):
    path = _state_path(day)
    if os.path.exists(path):
        os.remove(path)
    return jsonify(_build_fresh_day_state(day))


@app.route("/api/day/<int:day>/export_tracking", methods=["POST"])
def api_export_tracking(day: int):
    """인원 추적 트리들을, 가지(리프)마다 그날 시간표 한 줄로 정리한
    엑셀 파일을 만들어 내려준다. 화면에 떠 있는(아직 저장 안 됐을 수도
    있는) 최신 상태를 그대로 요청 본문으로 받아서 쓴다(디스크에 저장된
    파일을 다시 읽지 않음 - 저장 디바운스 타이밍과 무관하게 항상 지금
    보고 있는 화면 그대로 내보내기 위함)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = request.get_json(force=True)
    rows = _reconstruct_leaf_timetables(data)
    if not rows:
        return jsonify({"error": "추적 중인 그룹이 없습니다."}), 400

    slot_labels = data.get("slot_labels") or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{day}일차 인원추적"[:31]  # 엑셀 시트명 31자 제한

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    idle_fill = PatternFill("solid", fgColor="F2F2F2")

    headers = ["추적 그룹", "가지", "인원수"] + list(slot_labels)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 같은 트리(그룹) 안에서 가지 번호(1, 2, 3...)를 순서대로 매긴다.
    tree_branch_no: dict[str, int] = {}
    r = 2
    for row in rows:
        tree_label = row["tree_label"]
        tree_branch_no[tree_label] = tree_branch_no.get(tree_label, 0) + 1
        branch_no = tree_branch_no[tree_label]

        ws.cell(row=r, column=1, value=tree_label)
        ws.cell(row=r, column=2, value=f"가지 {branch_no}")
        ws.cell(row=r, column=3, value=row["count"])
        for si in range(len(slot_labels)):
            line_id = row["slot_line"].get(si)
            if line_id is None:
                continue
            cell = ws.cell(row=r, column=4 + si)
            cell.alignment = Alignment(horizontal="center")
            if line_id == "__idle__":
                cell.value = "미배치/휴식"
                cell.fill = idle_fill
            else:
                cell.value = line_id
                cell.fill = PatternFill("solid", fgColor=_line_fill_hex(line_id))
        r += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    for c in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "D2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"day{day}_tracking.xlsx",
    )


def main():
    parser = argparse.ArgumentParser(description="30일 계획 결과를 놓고 하루치 인원 배치를 수작업으로 조정하는 로컬 웹 도구")
    parser.add_argument("--dir", default=None, help="line_schedule.csv/daily_workforce.csv가 있는 디렉터리")
    parser.add_argument("--port", type=int, default=5000)
    parser.add_argument("--no-browser", action="store_true", help="자동으로 브라우저를 열지 않음")
    parser.add_argument(
        "--password", default=_DEFAULT_PASSWORD,
        help=f"접속 시 HTTP Basic Auth로 요구할 비밀번호(기본값: '{_DEFAULT_PASSWORD}', ngrok 등으로 "
             "외부 공유할 때 매번 --password를 안 줘도 자동 적용됨)",
    )
    parser.add_argument(
        "--no-password", action="store_true",
        help="비밀번호 보호를 끄고 실행(로컬(127.0.0.1)에서만 쓸 때)",
    )
    add_source_args(parser)
    args = parser.parse_args()

    global _PASSWORD
    _PASSWORD = None if args.no_password else args.password

    print("[정보] 라인/주문 데이터 로딩 중(필요인원 계산용, 시간이 좀 걸릴 수 있습니다)...")
    # resolve_source()의 --real-plan 기본 디렉터리는 "script_dir/real_plan"으로
    # 계산되는데, plot_day.py 등은 output/ 안에 있어서 output/real_plan이
    # 맞지만 이 파일은 opt_modeling_proto/ 바로 밑에 있으므로 output/을
    # 직접 붙여줘야 같은 output/real_plan 경로가 나온다.
    lines, orders, default_dir = resolve_source(args, os.path.join(_HERE, "output"))
    args.dir = args.dir or default_dir
    workers_lookup = build_workers_lookup(lines, orders)

    schedule_path = os.path.join(args.dir, "line_schedule.csv")
    workforce_path = os.path.join(args.dir, "daily_workforce.csv")
    with open(schedule_path, "rb") as f:
        schedule_bytes = f.read()
    schedule_fingerprint = hashlib.sha256(schedule_bytes).hexdigest()
    schedule_df = pd.read_csv(io.BytesIO(schedule_bytes))
    workforce_df = pd.read_csv(workforce_path)

    all_lines = sorted(schedule_df["line_id"].unique().tolist(), key=_line_sort_key)

    STATE["schedule_df"] = schedule_df
    STATE["workforce_df"] = workforce_df
    STATE["workers_lookup"] = workers_lookup
    STATE["all_lines"] = all_lines
    STATE["state_dir"] = os.path.join(args.dir, "staffing_state")
    STATE["schedule_fingerprint"] = schedule_fingerprint

    print(f"[정보] {schedule_path} 로드 완료 (라인 {len(all_lines)}개, {schedule_df['day'].nunique()}일치)")
    print(f"[정보] 편집 상태 저장 위치: {STATE['state_dir']}")
    print(f"[정보] http://127.0.0.1:{args.port} 에서 접속하세요")
    if _PASSWORD:
        print("[정보] 비밀번호 보호가 켜져 있습니다(아이디는 아무거나, 비밀번호만 확인).")

    if not args.no_browser:
        import threading
        import webbrowser
        threading.Timer(1.0, lambda: webbrowser.open(f"http://127.0.0.1:{args.port}")).start()

    app.run(host="127.0.0.1", port=args.port, debug=False)


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
data_pipeline/orders_from_excel.py

"수주진행현황" 엑셀(월별 시트에 수주 1건 = 1행)을 읽어서
scheduling.models.Order 목록으로 변환한다.

대상 행 필터(load_orders_from_excel의 기본 동작):
  (기준일 기준 납기가 남아있음 OR 납기가 ASAP) AND (상태가 완료/취소 상태가 아님)
  완료/취소 상태 = {"생산완료", "출고완료", "수주취소", "취소됨"} (EXCLUDED_STATUSES 참고)

엑셀 구조상 주의할 점(직접 파일을 열어서 확인한 내용):
  - "수주번호"/"no." 열은 전체 행의 약 21%가 비어 있어서 그 자체로는
    고유 식별자로 못 쓴다. "품번"(제품코드)도 여러 행이 공유한다(같은
    제품을 여러 차수에 나눠 수주하는 경우가 흔함) - 그래서 Order.order_id는
    이 파일에서 별도로 만들어 붙인다(아래 _make_order_id 참고).
  - 시트마다 열 구성이 미묘하게 다르다(예: 1월 시트엔 "사양" 열이 아예
    없어서 그 뒤 모든 열이 한 칸씩 밀려 있음, 월별 누적 통계 열 개수도
    시트마다 다름). 그래서 열 번호를 고정하지 않고, 매 시트마다 2행
    헤더를 다시 읽어서 열 이름으로 위치를 찾는다(_resolve_columns).
  - "납기(인폼 기준)"/"조정납기일" 두 열 다 날짜가 아닌 텍스트가 섞여
    있다("ASAP", "4월중", "2월초", "25-01차"(발주 차수를 잘못 옮겨 적은
    것으로 보임), "수주등록x", "-" 등). 이 파일은 datetime과 "ASAP"
    문자열만 명확히 해석하고, 그 외 텍스트는 "해석 불가"로 분류해서
    기본적으로 제외 + 경고 로그를 남긴다(조용히 잘못 끼워맞추지 않음).
  - "생산완료" 열은 상태 플래그가 아니라 "지금까지 생산한 수량"이다
    (헤더 이름이 오해하기 쉬움). Order.quantity로는 "생산잔량"(아직
    안 만든 수량)을 쓴다 - 스케줄러는 오늘부터 앞으로 계획을 짜는
    것이므로 이미 만든 수량까지 포함한 전체 수주수량을 넣으면 안 된다.
"""

from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.utils.datetime import from_excel

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # opt_modeling_proto/ 를 import 경로에 추가
from scheduling.models import Order

HEADER_ROW = 5
DATA_START_ROW = 6

# 엑셀상 납기 그대로가 아니라 이만큼 앞당긴 날짜를 프로그램상 마감일로
# 쓴다 - 실제로는 납기 당일이 아니라 그보다 며칠 먼저 생산을 끝내둬야
# 완제품검사가 가능하기 때문. '완제품검사' 열이 'y'인 제품은 검사가
# 하루면 끝나므로 PRODUCTION_LEAD_DAYS_INSPECTED, 아니면 기존처럼
# PRODUCTION_LEAD_DAYS_DEFAULT를 쓴다(제품별로 다름 - 아래 루프에서
# 매 행마다 고른다). 앞당긴 결과가 1일차(기준일 당일)보다 이르게
# 나오면(=엑셀 납기가 이미 코앞이라 여유가 그만큼도 안 남음) 1일차로
# clamp한다 - 0/음수 day는 스케줄러가 다루는 day index 체계(1-indexed)에
# 없는 값이라 그대로 넘기면 에러가 난다.
PRODUCTION_LEAD_DAYS_DEFAULT = 6
PRODUCTION_LEAD_DAYS_INSPECTED = 1

# 부자재/원료 입고예정일에 더하는 여유일. 입고 당일 바로 생산 가능한 게
# 아니라 부자재는 검수 등으로 +SUBMATERIAL_LEAD_DAYS일, 원료는 내용물
# 생산+검사 시간으로 며칠이 지나야 실제로 투입 가능하다는 현장 기준.
# '내용물검사' 열이 'y'인 제품은 검사가 하루면 끝나므로
# RAW_MATERIAL_LEAD_DAYS_INSPECTED, 아니면 기존처럼
# RAW_MATERIAL_LEAD_DAYS_DEFAULT를 쓴다(제품별로 다름 - 아래 루프에서
# 매 행마다 고른다). earliest_start_day는 "(부자재입고예정일 +
# SUBMATERIAL_LEAD_DAYS) vs (원료입고예정일 + 원료 lead days)" 중 더
# 늦은 날짜로 정한다(입고일 자체를 그냥 비교하는 게 아님 - 각각 여유일을
# 더한 "실제 투입 가능일"끼리 비교).
SUBMATERIAL_LEAD_DAYS = 1
RAW_MATERIAL_LEAD_DAYS_DEFAULT = 7
RAW_MATERIAL_LEAD_DAYS_INSPECTED = 1

# 논리 필드 이름 -> 엑셀 헤더 텍스트(공백/줄바꿈 제거 후 비교, _normalize_header 참고).
# 시트마다 실제 열 번호가 다르므로 이름으로 찾는다.
COLUMN_ALIASES: dict[str, list[str]] = {
    "order_no": ["수주번호"],
    "category": ["제품군"],
    "vendor": ["발주처"],
    "product_id": ["품번"],
    "product_name": ["품명"],
    "unit": ["단위"],
    "due_date": ["납기(인폼기준)"],
    "adjusted_due_date": ["조정납기일"],
    "order_qty": ["수주수량"],
    "produced_qty": ["생산완료"],
    "remaining_qty": ["생산잔량"],
    "submaterial_date": ["부자재입고예정일"],
    "raw_material_date": ["원료입고예정일"],
    "status": ["상태"],
    "content_inspection": ["내용물검사"],
    "finished_inspection": ["완제품검사"],
}

# 이미 끝났거나(생산/출고 완료) 더 이상 유효하지 않은(취소) 주문 상태.
# "상태" 열은 자유 텍스트라 이 외에도 진행 메모성 문구가 잔뜩 섞여
# 있지만(예: "용기 9/12, 단상자 9/11..."), 그런 진행중 메모는 스케줄링
# 대상에서 뺄 이유가 없으므로 정확히 이 값들과 일치할 때만 제외한다.
EXCLUDED_STATUSES = {"생산완료", "출고완료", "수주취소", "취소됨"}

# 제품(행)별로 호환되는 라인타입마다 rpm/시간당 Capa(rate)/동시 투입인원(workers)을
# 적어두는 5개 컬럼 그룹(셀라인/단발/10열기/로타리/튜브라인 순서, 각 3열씩
# T~AH). --read-specs excel 모드에서만 쓰인다(plan_from_orders.py 참고).
#
# COLUMN_ALIASES처럼 헤더 텍스트로 못 찾고 고정 열 번호로 찾는 이유: 이
# 5개 그룹의 헤더가 전부 "rpm"/"시간당 Capa"/"동시 작업인원 수(명)"로
# 동일해서(엑셀 원본이 실제로 이렇게 돼 있음) _resolve_columns의 "헤더
# 텍스트 -> 열 번호" 방식으로는 그룹을 구분할 수 없다 - 5번 다 같은 키로
# 덮어써져서 마지막(튜브라인) 것만 남아버림. 그래서 이 5개 그룹만
# 예외적으로 위치 고정 방식을 쓴다(사용자가 직접 확인해서 알려준 위치:
# T/U/V, W/X/Y, Z/AA/AB, AC/AD/AE, AF/AG/AH). 시트 레이아웃이 바뀌면
# 조용히 엉뚱한 값을 읽을 수 있으므로, load_orders_from_excel이 매번
# 헤더 텍스트가 예상과 맞는지 확인해서 안 맞으면 경고를 남긴다.
LINE_SPEC_COLUMNS: dict[str, tuple[int, int, int]] = {
    # line_type_id: (rpm 열, 시간당Capa/rate 열, 동시투입인원/workers 열) - 1-indexed
    "셀라인": (20, 21, 22),      # T, U, V
    "단발": (23, 24, 25),        # W, X, Y
    "10열기": (26, 27, 28),      # Z, AA, AB
    "로타리": (29, 30, 31),      # AC, AD, AE
    "튜브라인": (32, 33, 34),    # AF, AG, AH
}


def _normalize_header(text) -> str:
    """헤더 셀 값에서 공백/줄바꿈을 전부 제거해서 비교용 키로 만든다.
    "납기\n(인폼 기준)"처럼 셀 안에 줄바꿈이 들어있는 헤더가 있어서
    단순 문자열 비교로는 안 걸리기 때문."""
    return re.sub(r"\s+", "", str(text)) if text is not None else ""


def _resolve_columns(ws) -> dict[str, int | None]:
    """이 시트의 2행 헤더를 읽어서 {논리필드명: 열번호} 매핑을 만든다.
    시트에 그 헤더가 없으면 None(해당 시트에서는 그 필드를 못 읽는다는 뜻)."""
    header_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _normalize_header(ws.cell(row=HEADER_ROW, column=c).value)
        if key:
            header_to_col[key] = c

    resolved: dict[str, int | None] = {}
    for field, aliases in COLUMN_ALIASES.items():
        resolved[field] = next(
            (header_to_col[a] for a in aliases if a in header_to_col), None
        )
    return resolved


@dataclass
class DeadlineResolution:
    kind: str  # "date" / "asap" / "unresolved"
    deadline_date: date | None = None
    detail: str | None = None  # kind=="unresolved"일 때 원본 값 설명(경고 로그용)


def _excel_serial_to_date(value: int | float) -> date | None:
    """날짜 열의 셀 서식이 (원래는 날짜여야 하는데) 회계/숫자 서식 등으로
    잘못 저장돼서 openpyxl이 datetime 대신 raw 엑셀 시리얼 넘버(정수/실수)를
    그대로 돌려주는 경우가 실제로 있었다(예: due_date 셀이
    '_(* #,##0_)...' 회계 서식이라 46294 같은 숫자로 나옴 - 변환해보면
    2026-09-29처럼 멀쩡한 날짜). 서식과 무관하게 값 자체는 항상 신뢰할 수
    있으므로, 날짜로 변환 가능한 숫자면 여기서 구제한다. 음수/터무니없이
    큰 값처럼 애초에 날짜가 아닌 숫자가 섞여 들어온 경우에는 from_excel이
    예외를 던지므로 그때만 None(해석 불가)으로 처리한다."""
    try:
        return from_excel(value).date()
    except (ValueError, OverflowError, TypeError):
        return None


def _resolve_single_deadline_cell(value) -> DeadlineResolution:
    if isinstance(value, datetime):
        return DeadlineResolution(kind="date", deadline_date=value.date())
    if isinstance(value, date):  # openpyxl이 시간 없는 날짜를 date로 줄 수도 있음
        return DeadlineResolution(kind="date", deadline_date=value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        converted = _excel_serial_to_date(value)
        if converted is not None:
            return DeadlineResolution(kind="date", deadline_date=converted)
        return DeadlineResolution(kind="unresolved", detail=f"날짜로 변환 안 되는 숫자: {value!r}")
    if isinstance(value, str):
        s = value.strip()
        if s.lower() == "asap":
            return DeadlineResolution(kind="asap")
        if s and s != "-":
            return DeadlineResolution(kind="unresolved", detail=f"해석 불가 텍스트: {s!r}")
    return DeadlineResolution(kind="unresolved", detail="비어 있음")


def _resolve_deadline(adjusted_raw, due_raw) -> DeadlineResolution:
    """조정납기일을 우선 시도하고(날짜든 ASAP든 해석되면 그걸 쓴다),
    조정납기일이 해석 안 되면(비어있거나 "25-01차" 같은 발주차수 오기입
    등) 원래 납기(인폼 기준)로 폴백한다. 둘 다 해석 안 되면 unresolved."""
    adjusted = _resolve_single_deadline_cell(adjusted_raw)
    if adjusted.kind in ("date", "asap"):
        return adjusted
    due = _resolve_single_deadline_cell(due_raw)
    if due.kind in ("date", "asap"):
        return due
    return DeadlineResolution(
        kind="unresolved",
        detail=f"조정납기일({adjusted.detail}) / 납기({due.detail}) 둘 다 해석 불가",
    )


def _to_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _extract_line_specs(ws, row: int) -> tuple[dict[str, float], dict[str, int]]:
    """LINE_SPEC_COLUMNS(T~AH)에서 이 행(제품)이 실제로 호환되는
    라인타입별 rate(시간당 생산량)/workers(필요인원)를 읽는다. 시간당
    Capa가 비어 있거나 0 이하면 그 라인타입은 호환 안 되는 것으로 보고
    dict에 아예 안 넣는다(Order.compatible_line_types()가 rate>0인
    항목만 세므로 자연히 빠짐)."""
    rate: dict[str, float] = {}
    workers: dict[str, int] = {}
    for line_type_id, (_rpm_col, capa_col, workers_col) in LINE_SPEC_COLUMNS.items():
        capa = _to_number(ws.cell(row=row, column=capa_col).value)
        if capa is None or capa <= 0:
            continue
        rate[line_type_id] = capa
        workers[line_type_id] = int(_to_number(ws.cell(row=row, column=workers_col).value) or 0)
    return rate, workers


def _check_line_spec_headers(ws) -> list[str]:
    """LINE_SPEC_COLUMNS가 가리키는 고정 열 위치의 헤더 텍스트가 실제로
    rpm/시간당Capa/동시투입인원인지 확인한다(위치 고정 방식이라 시트
    레이아웃이 바뀌면 조용히 엉뚱한 값을 읽을 위험이 있어서, 매번 확인
    후 안 맞으면 경고 문자열 목록을 돌려준다)."""
    problems = []
    for line_type_id, (rpm_col, capa_col, workers_col) in LINE_SPEC_COLUMNS.items():
        rpm_header = _normalize_header(ws.cell(row=HEADER_ROW, column=rpm_col).value)
        capa_header = _normalize_header(ws.cell(row=HEADER_ROW, column=capa_col).value)
        workers_header = _normalize_header(ws.cell(row=HEADER_ROW, column=workers_col).value)
        if "rpm" not in rpm_header.lower():
            problems.append(f"{line_type_id}: rpm 열({rpm_col}) 헤더가 예상과 다름({rpm_header!r})")
        if "Capa" not in capa_header:
            problems.append(f"{line_type_id}: 시간당Capa 열({capa_col}) 헤더가 예상과 다름({capa_header!r})")
        if "작업인원" not in workers_header:
            problems.append(f"{line_type_id}: 동시투입인원 열({workers_col}) 헤더가 예상과 다름({workers_header!r})")
    return problems


# "5매입", "(10매)", "10+1매/자급"처럼 품명에 붙는 매수 표기를 찾는다.
# "+"로 여러 숫자가 이어진 경우(예: "10+1매") 그 합을 매수로 본다.
_SHEET_COUNT_RE = re.compile(r"(\d+(?:\+\d+)*)\s*매")


def _extract_sheet_count(product_name: str) -> int | None:
    """품명에서 "(숫자)매" 패턴을 찾아 매수(시트 수)를 반환한다. 못
    찾으면 None. 찾은 숫자(+가 있으면 합산한 값)가 1이면 "매수 표기가
    실질적으로 없는 것"과 같으므로(예: "1매입"은 그냥 낱개 제품)
    None을 반환해서 호출하는 쪽에서 별도 처리 없이 지나가게 한다."""
    if not product_name:
        return None
    m = _SHEET_COUNT_RE.search(product_name)
    if not m:
        return None
    total = sum(int(part) for part in m.group(1).split("+"))
    return total if total != 1 else None


def _is_marked_yes(value) -> bool:
    """'내용물검사'/'완제품검사' 열이 'y'로 체크돼 있는지 확인한다
    (대소문자/앞뒤 공백 무시). 비어 있거나 다른 값이면 False."""
    return value is not None and str(value).strip().lower() == "y"


def _resolve_supply_date(value) -> date | None:
    """'부자재입고예정일'/'원료입고예정일' 열은 날짜값이 기본이지만,
    예전 '부자재' 통합 열처럼 "완료"/"확인중"/"재고사용"/"5/12~5/20"
    같은 자유 텍스트가 섞여 들어올 가능성을 배제할 수 없다. 텍스트는
    날짜로 신뢰성 있게 해석할 방법이 없으므로(완료/재고사용처럼 "이미
    준비됨"을 뜻하는 경우도 있고, 애매한 부분 날짜 표기도 있음) 명확한
    날짜 값일 때만 해석하고, 그 외는 전부 "제약 없음"으로 취급한다 -
    조용히 잘못 해석해서 불필요한 생산불가 제약을 거는 것보다 안전한
    쪽을 택함. 숫자(raw 엑셀 시리얼 넘버)도 날짜로 인정한다 - 셀 서식이
    날짜가 아니게 저장된 경우의 fallback(_excel_serial_to_date 참고)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _excel_serial_to_date(value)
    return None


def load_orders_from_excel(
    path: str,
    reference_date: date | None = None,
    verbose: bool = True,
    read_specs_from_excel: bool = False,
) -> tuple[list[Order], dict]:
    """수주진행현황 엑셀을 읽어서 필터링된 (Order 목록, 요약 통계 dict)를 반환한다.

    stats 키: scanned/excluded_status/excluded_deadline_passed/
    excluded_deadline_unresolved/excluded_nonpositive_qty/
    excluded_deadline_before_lead/included
    (plan_from_orders.py의 plan_report 엑셀에서 제외 사유별 개수를
    보여줄 때 씀 - filter_and_attach_rates의 2차 필터 stats와 합쳐서
    쓰인다).

    필터: (reference_date 기준 납기가 아직 안 지났음 OR 납기가 ASAP)
          AND (상태가 EXCLUDED_STATUSES에 없음)
          AND (생산 lead time을 뺀 실제 생산 마감일이 1일차 이상)

    reference_date를 안 주면 오늘 날짜를 쓴다. Order.deadline_day는
    "reference_date를 1일차로 하는 스케줄링 horizon 안에서 며칠째인지"로
    변환해서 넣는다(reference_date 당일이 마감이면 deadline_day=1) -
    scheduling.models.Order.deadline_day의 정의(1-indexed) 그대로.
    엑셀 납기 자체는 아직 안 지났더라도, 거기서 생산 lead time을 뺀
    실제 생산 마감일이 1일차보다 이르면(=이미 물리적으로 시간이 없는
    주문) 1일차로 눙쳐 넘기지 않고 아예 제외한다(excluded_deadline_before_lead).

    rate/workers: read_specs_from_excel=False(기본)면 이 시점엔 알 수
    없으므로 빈 dict로 두고, 호출하는 쪽(plan_from_orders.py의
    filter_and_attach_rates)이 제품군(category) 기준으로 채워 넣는다.
    read_specs_from_excel=True면 여기서 바로 LINE_SPEC_COLUMNS(T~AH,
    라인타입별 rpm/시간당Capa/동시투입인원)를 읽어서 채운다 - 제품군이
    아니라 그 행(제품) 자체에 적힌 실측값을 쓰는 방식. 어느 쪽을 쓸지는
    plan_from_orders.py의 --read-specs 플래그로 고른다.

    category 보정 + 멀티시트 처리: '제품군'이 "파우치"면 "마스크"로
    바꾼다. 그 결과 category가 "마스크"인 행은 '품명'에서 "(숫자)매"
    표기(예: "5매입", "10+1매")를 찾아서(_extract_sheet_count 참고),
    있으면 '단위'를 확인한다 - "SET"이면 생산잔량(개수 단위로 환산해야
    스케줄러의 rate와 맞음)에 매수를 곱하고, "EA"면 생산잔량은 그대로
    두되 category를 "마스크_멀티시트"로 바꿔서 별도 라인 스펙
    (plan_from_orders.py의 CATEGORY_LINE_SPECS)이 적용되게 한다.

    deadline_day: 엑셀 납기 그대로가 아니라 며칠 앞당긴 날짜를 쓴다(실제
    출고를 위해 납기보다 먼저 생산을 끝내둬야 완제품검사가 가능해서) -
    앞당기는 일수는 '완제품검사' 열이 'y'면 PRODUCTION_LEAD_DAYS_INSPECTED,
    아니면 PRODUCTION_LEAD_DAYS_DEFAULT(제품마다 다름). 앞당긴 값이
    1보다 작아지면(=실제 생산 마감일이 이미 지남) 1로 clamp하지 않고
    그 주문 자체를 제외한다(excluded_deadline_before_lead).

    earliest_start_day: '부자재입고예정일'/'원료입고예정일' 두 열 중 실제
    날짜 값이 있는 쪽을 각각 해석하고(텍스트는 자유 서식이 섞여 있어
    무시함, _resolve_supply_date 참고), 각각에 여유일을 더한 "실제 투입
    가능일"을 구한 다음, 그 둘 중 더 늦은 날짜를 "이 주문의 생산이
    가능해지는 첫날"로 넣는다(입고예정일 자체를 그냥 비교하는 게 아니라
    여유일을 더한 값끼리 비교 - 부자재/원료 둘 다 들어와야 생산을 시작할
    수 있으므로). 부자재 쪽 여유일은 항상 SUBMATERIAL_LEAD_DAYS, 원료
    쪽은 '내용물검사' 열이 'y'면 RAW_MATERIAL_LEAD_DAYS_INSPECTED, 아니면
    RAW_MATERIAL_LEAD_DAYS_DEFAULT(제품마다 다름). reference_date
    당일이면 곧바로 1일차부터 가능하므로 제약 없음과 동일 - None으로 둠.
    계획기간(horizon) 초과 여부는 이 함수가 horizon 길이를 모르므로
    여기서는 확인하지 않고, 호출하는 쪽(plan_from_orders.py의
    filter_and_attach_rates)에서 처리한다.
    """
    reference_date = reference_date or date.today()
    # 시트가 "YYYY년MM월" 형식으로 매달 하나씩 쌓이는 구조인데, 과거 달
    # 시트는 이미 다 지난 데이터라 다시 읽을 필요가 없다("납기지남"으로
    # 걸러지긴 하지만 시간 낭비) - reference_date가 속한 달의 시트 하나만
    # 읽는다. 다음 달이 되면 자동으로 그 달 시트를 찾게 된다.
    target_sheet_name = f"{reference_date.year}년{reference_date.month:02d}월"

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if target_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"'{target_sheet_name}' 시트를 찾을 수 없습니다. "
            f"엑셀에 있는 시트: {wb.sheetnames}"
        )
    sheet_names = [target_sheet_name]

    orders: list[Order] = []
    seen_order_ids: dict[str, int] = {}  # order_id -> 지금까지 나온 횟수(충돌 시 접미사 붙이는 용도)

    # 요약 카운터 + 경고 상세(사람이 나중에 엑셀에서 직접 확인할 수 있게
    # 시트/행/품번을 같이 남긴다).
    stats = {
        "scanned": 0,
        "excluded_status": 0,
        "excluded_deadline_passed": 0,
        "excluded_deadline_unresolved": 0,
        "excluded_nonpositive_qty": 0,
        "excluded_deadline_before_lead": 0,
        "included": 0,
    }
    warnings: list[str] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        cols = _resolve_columns(ws)
        missing = [f for f, c in cols.items() if c is None]
        if missing:
            warnings.append(f"[{sheet_name}] 헤더에서 못 찾은 필드(이 시트에서는 해당 값 없음 취급): {missing}")
        if read_specs_from_excel:
            header_problems = _check_line_spec_headers(ws)
            if header_problems:
                warnings.append(
                    f"[{sheet_name}] --read-specs excel 모드인데 라인별 rpm/시간당Capa/동시투입인원 "
                    f"열 위치가 예상과 다릅니다(시트 레이아웃 변경 의심, LINE_SPEC_COLUMNS 확인 필요): "
                    + "; ".join(header_problems)
                )

        def cell(field: str, row: int):
            c = cols[field]
            return ws.cell(row=row, column=c).value if c is not None else None

        for r in range(DATA_START_ROW, ws.max_row + 1):
            product_id = cell("product_id", r)
            if product_id is None or str(product_id).strip() == "":
                continue  # 품번 없는 행은 진짜 데이터가 아닌 빈 줄/서식용 행으로 취급
            product_id = str(product_id).strip()
            stats["scanned"] += 1
            row_ref = f"{sheet_name}!row{r}({product_id})"

            # ---- 상태 필터 ----
            status = cell("status", r)
            status_text = str(status).strip() if status is not None else ""
            if status_text in EXCLUDED_STATUSES:
                stats["excluded_status"] += 1
                continue

            # ---- 납기 필터 ----
            deadline = _resolve_deadline(cell("adjusted_due_date", r), cell("due_date", r))
            if deadline.kind == "unresolved":
                stats["excluded_deadline_unresolved"] += 1
                warnings.append(f"{row_ref}: 납기 {deadline.detail} - 제외")
                continue
            if deadline.kind == "date" and deadline.deadline_date < reference_date:
                stats["excluded_deadline_passed"] += 1
                continue

            # ---- category 보정(파우치->마스크) + 멀티시트(매수) 처리 ----
            category_raw = cell("category", r)
            category = str(category_raw).strip() if category_raw is not None else ""
            if category == "파우치":
                category = "마스크"

            sheet_multiplier = 1
            if category == "마스크":
                sheet_count = _extract_sheet_count(str(cell("product_name", r) or ""))
                if sheet_count is not None:
                    unit_raw = cell("unit", r)
                    unit = str(unit_raw).strip().upper() if unit_raw is not None else ""
                    if unit == "SET":
                        sheet_multiplier = sheet_count
                    elif unit == "EA":
                        category = "마스크_멀티시트"
                    else:
                        warnings.append(
                            f"{row_ref}: 품명에 '{sheet_count}매' 표기가 있으나 단위가 SET/EA가 "
                            f"아님({unit!r}) - 매수 처리 생략"
                        )

            # ---- 수량: 생산잔량(아직 안 만든 수량) 우선, 없으면 수주수량-생산완료로 계산 ----
            remaining_qty = _to_number(cell("remaining_qty", r))
            if remaining_qty is None:
                order_qty = _to_number(cell("order_qty", r))
                produced_qty = _to_number(cell("produced_qty", r))
                if order_qty is not None and produced_qty is not None:
                    remaining_qty = order_qty - produced_qty
                    warnings.append(f"{row_ref}: 생산잔량 열이 비어 있어 수주수량-생산완료로 계산함 ({remaining_qty:g})")
                else:
                    remaining_qty = order_qty  # 그마저 없으면 수주수량 그대로(최소한의 폴백)
                    warnings.append(f"{row_ref}: 생산잔량/생산완료 모두 확인 불가, 수주수량을 그대로 사용")
            if remaining_qty is None or remaining_qty <= 0:
                stats["excluded_nonpositive_qty"] += 1
                continue
            if sheet_multiplier != 1:
                remaining_qty = remaining_qty * sheet_multiplier

            # ---- 제품별 lead days 결정 ('완제품검사'/'내용물검사' 열이 'y'면 단축) ----
            is_finished_inspected = _is_marked_yes(cell("finished_inspection", r))
            is_content_inspected = _is_marked_yes(cell("content_inspection", r))
            production_lead_days = (
                PRODUCTION_LEAD_DAYS_INSPECTED if is_finished_inspected else PRODUCTION_LEAD_DAYS_DEFAULT
            )
            raw_material_lead_days = (
                RAW_MATERIAL_LEAD_DAYS_INSPECTED if is_content_inspected else RAW_MATERIAL_LEAD_DAYS_DEFAULT
            )

            # ---- deadline_day 계산 (엑셀 납기 - production_lead_days일) ----
            # 엑셀 납기 자체는 아직 안 지났어도(위 "납기 필터"를 통과했어도),
            # 완제품검사 등을 위해 며칠 앞당겨 생산을 끝내야 하는 실제
            # 생산 마감일(=deadline_day)이 오늘(1일차)보다 이르게 나올 수
            # 있다 - 즉 물리적으로 이미 늦은 주문. 예전에는 이 경우 그냥
            # 1일차로 clamp해서 "오늘이 마감"인 것처럼 취급했지만, 그러면
            # 실제로는 생산할 시간이 없는 주문이 마치 아직 여유가 있는
            # 것처럼 보여서 문제였다 - clamp하지 않고 아예 제외한다.
            deadline_day: int | None
            if deadline.kind == "asap":
                deadline_day = None
            else:
                raw_deadline_day = (deadline.deadline_date - reference_date).days + 1
                deadline_day = raw_deadline_day - production_lead_days
                if deadline_day < 1:
                    stats["excluded_deadline_before_lead"] += 1
                    warnings.append(
                        f"{row_ref}: 납기({deadline.deadline_date.isoformat()})에서 생산 lead time"
                        f"({production_lead_days}일)을 빼면 이미 지났음(day {deadline_day}) - 제외"
                    )
                    continue

            # ---- earliest_start_day 계산 ----
            # 부자재/원료 둘 다 도착해야 생산을 시작할 수 있으므로, 각각
            # 입고예정일에 여유일(SUBMATERIAL_LEAD_DAYS/raw_material_lead_days)을
            # 더한 "실제 투입 가능일"을 구한 뒤 그 중 더 늦은 날짜를 쓴다
            # (입고예정일 자체를 그냥 비교하는 게 아님).
            ready_dates = []
            submaterial_date = _resolve_supply_date(cell("submaterial_date", r))
            if submaterial_date is not None:
                ready_dates.append(submaterial_date + timedelta(days=SUBMATERIAL_LEAD_DAYS))
            raw_material_date = _resolve_supply_date(cell("raw_material_date", r))
            if raw_material_date is not None:
                ready_dates.append(raw_material_date + timedelta(days=raw_material_lead_days))

            earliest_start_day: int | None = None
            if ready_dates:
                latest_ready_date = max(ready_dates)
                candidate = (latest_ready_date - reference_date).days + 1
                if candidate > 1:  # 1일차(오늘)부터 가능하면 사실상 제약 없음과 동일
                    earliest_start_day = candidate

            order_id = _make_order_id(cell("order_no", r), sheet_name, r, product_id, seen_order_ids)

            vendor_raw = cell("vendor", r)
            product_name_raw = cell("product_name", r)

            if read_specs_from_excel:
                rate, workers = _extract_line_specs(ws, r)
            else:
                rate, workers = {}, {}

            orders.append(
                Order(
                    order_id=order_id,
                    product_id=product_id,
                    category=category,
                    quantity=int(round(remaining_qty)),
                    product_name=str(product_name_raw).strip() if product_name_raw is not None else "",
                    vendor=str(vendor_raw).strip() if vendor_raw is not None else "",
                    deadline_day=deadline_day,
                    earliest_start_day=earliest_start_day,
                    rate=rate,
                    workers=workers,
                    content_inspection=is_content_inspected,
                    finished_inspection=is_finished_inspected,
                    submaterial_date=submaterial_date,
                    raw_material_date=raw_material_date,
                    raw_deadline_date=deadline.deadline_date if deadline.kind == "date" else None,
                )
            )
            stats["included"] += 1

    if verbose:
        print(f"[정보] 기준일: {reference_date.isoformat()}, 대상 시트: {sheet_names}")
        print(
            f"[정보] 스캔 {stats['scanned']}건 -> 포함 {stats['included']}건 "
            f"(제외: 완료/취소상태 {stats['excluded_status']}, 납기지남 {stats['excluded_deadline_passed']}, "
            f"납기해석불가 {stats['excluded_deadline_unresolved']}, 잔량0이하 {stats['excluded_nonpositive_qty']}, "
            f"생산lead time반영후납기지남 {stats['excluded_deadline_before_lead']})"
        )
        if warnings:
            print(f"[경고] {len(warnings)}건의 주의사항(수동 확인 권장):")
            for w in warnings[:50]:
                print(f"  - {w}")
            if len(warnings) > 50:
                print(f"  ... 외 {len(warnings) - 50}건 더")

    return orders, stats


def _make_order_id(
    order_no,
    sheet_name: str,
    row: int,
    product_id: str,
    seen: dict[str, int],
) -> str:
    """수주번호가 있으면 그걸 order_id로 쓰고(전체 파일에서 21% 정도가
    비어 있어서 기본값을 대비해야 함), 없으면 시트명+행번호+품번으로
    만든다(사람이 나중에 엑셀에서 원본 행을 다시 찾기 쉽게). 어느 쪽이든
    같은 값이 중복 발생하면(수주번호가 여러 행에 재사용된 경우 등)
    "_2", "_3" 접미사를 붙여서 전체적으로 고유하게 만든다.
    """
    if order_no is not None and str(order_no).strip():
        base = str(order_no).strip()
    else:
        base = f"{sheet_name}_r{row}_{product_id}"

    count = seen.get(base, 0) + 1
    seen[base] = count
    return base if count == 1 else f"{base}_{count}"


def save_orders_json(orders: list[Order], path: str) -> None:
    """order 목록을 JSON 배열로 저장한다. 각 원소는 Order 필드 그대로
    (order_id/product_id/category/quantity/deadline_day/earliest_start_day/
    backlog_cost_per_unit_per_day/rate/workers) - scheduling.example_data.load_data_from_json이 기대하는
    최종 데이터는 {"lines": [...], "orders": [...]}인데, 이 스크립트는
    아직 lines(라인별 rate/workers)를 모르므로 orders 배열만 저장해둔다.
    다음 단계(rate/worker 계산 스크립트)에서 lines를 채워 최종 JSON으로
    합친다.
    """
    import json
    from dataclasses import asdict

    with open(path, "w", encoding="utf-8") as f:
        json.dump([asdict(o) for o in orders], f, ensure_ascii=False, indent=2)


def save_orders_csv(orders: list[Order], path: str) -> None:
    """order 목록을 사람이 엑셀에서 바로 열어 검토하기 쉬운 CSV로 저장한다.
    rate/workers는 이 단계에선 항상 비어 있어서(다음 단계에서 채워짐)
    CSV에는 넣지 않는다 - deadline_day 대신 사람이 읽기 쉬운 형태로
    "ASAP" 또는 며칠째인지를 같이 보여준다.
    """
    import csv

    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["order_id", "product_id", "category", "quantity", "deadline_day", "earliest_start_day"])
        for o in orders:
            writer.writerow([
                o.order_id,
                o.product_id,
                o.category,
                o.quantity,
                "ASAP" if o.deadline_day is None else o.deadline_day,
                "" if o.earliest_start_day is None else o.earliest_start_day,
            ])


if __name__ == "__main__":
    import argparse

    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace", line_buffering=True)

    parser = argparse.ArgumentParser(description="수주진행현황 엑셀 -> Order 목록 변환")
    parser.add_argument("path", help="수주진행현황 엑셀 파일 경로")
    parser.add_argument("--reference-date", default=None, help="기준일(YYYY-MM-DD, 생략하면 오늘)")
    parser.add_argument(
        "--output-dir",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "output"),
        help="orders.json / orders.csv를 저장할 디렉터리",
    )
    parser.add_argument("--no-save", action="store_true", help="파일로 저장하지 않고 콘솔 미리보기만 출력")
    args = parser.parse_args()

    ref = date.fromisoformat(args.reference_date) if args.reference_date else None
    result, _stats = load_orders_from_excel(args.path, reference_date=ref)
    print(f"\n총 {len(result)}건의 Order 반환됨. 처음 5건:")
    for o in result[:5]:
        print(f"  {o.order_id} | {o.product_id} | {o.category} | qty={o.quantity} | deadline_day={o.deadline_day}")

    if not args.no_save:
        os.makedirs(args.output_dir, exist_ok=True)
        json_path = os.path.join(args.output_dir, "orders.json")
        csv_path = os.path.join(args.output_dir, "orders.csv")
        save_orders_json(result, json_path)
        save_orders_csv(result, csv_path)
        print(f"\n[저장 완료]\n  {json_path}\n  {csv_path}")

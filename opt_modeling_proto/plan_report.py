# -*- coding: utf-8 -*-
"""
plan_report.py

plan_from_orders.py가 CP-SAT으로 스케줄을 푼 직후에, 사람이 바로 열어볼
수 있는 결과 요약을 엑셀 한 장으로 만드는 write_plan_report_excel() 하나만
담은 모듈. scheduling/report.py(print_report/save_outputs/plot_gantt)와
따로 분리한 이유: 그쪽은 schedule_optimizer.py의 내장 예시 데이터에도
그대로 쓰이는 범용 리포트라 category 이름을 몰라도 되지만, 이 리포트는
plan_from_orders.py의 실제 수주 데이터 전용 category 이름("마스크"/
"마스크_멀티시트"/"용기"/"튜브")을 그대로 하드코딩해서 "마스크 환산 총
생산량" 같은 그 파이프라인만의 지표를 계산하므로 범용 모듈에 넣기엔
맞지 않는다.

plan_from_orders.py -> plan_report.py 방향으로만 import한다(반대 방향
없음) - 순환 import 걱정 없이 plan_from_orders.py에서 그냥 함수를
가져다 쓰면 된다.
"""

from __future__ import annotations

from datetime import date, timedelta

from scheduling.models import Line, Order, OVERTIME_LOCAL_SLOTS, ScheduleConfig, ScheduleResult, SLOTS_PER_DAY

# production_efficiency.py의 기본 WEIGHT_CONTAINER_TUBE와 같은 값 - 마스크
# 환산 총 생산량 = 마스크 + 마스크_멀티시트 + (용기+튜브)*이 값.
MASK_EQUIVALENT_WEIGHT = 5


def _build_workers_lookup(lines: list[Line], orders: list[Order]) -> dict[tuple[str, str], int]:
    """{(물리 line_id, order_id): 그 라인에서 이 주문 생산에 필요한 인원수}.
    output/_data_source.py에 이미 같은 함수(build_workers_lookup)가 있지만,
    그 모듈이 plan_from_orders를 import하므로 여기서 가져다 쓰면 순환
    import가 생긴다 - 그래서 scheduling.pooling으로 직접 다시 만든다
    (로직 자체는 완전히 동일)."""
    from scheduling.pooling import build_line_pools

    lookup: dict[tuple[str, str], int] = {}
    for pool in build_line_pools(lines, orders):
        for order_id, workers in pool.workers.items():
            for line_id in pool.line_ids:
                lookup[(line_id, order_id)] = workers
    return lookup


def write_plan_report_excel(
    result: ScheduleResult,
    orders: list[Order],
    lines: list[Line],
    config: ScheduleConfig,
    reference_date: date,
    load_stats: dict,
    filter_stats: dict,
    output_path: str,
    deadline_window_days: int,
) -> None:
    """plan_from_orders.py를 돌릴 때마다 사람이 바로 열어볼 결과 요약을
    엑셀 한 장(시트 하나, 섹션을 위에서 아래로 쌓는 구성)으로 만든다.

    섹션 구성(위->아래):
      1) 기준 정보 + 제외 사유별 개수(엑셀 로딩 단계의 load_stats와
         ramp/카테고리 필터 단계의 filter_stats를 합쳐서 보여줌 -
         "납기 해석 불가"/"납기 지남"은 load_stats, "납기 N일 초과"는
         filter_stats에서 옴, N=deadline_window_days)
      2) 주문별 상세: 발주처/제품명/납기일/생산량/생산잔량/최초생산일/
         생산완료일 (day index는 전부 reference_date 기준 실제 달력
         날짜로 변환 - "15일차"보다 "2026-07-24"가 바로 읽기 쉬움).
         생산잔량은 ASAP 주문이 backlog를 못 채웠을 때만 0보다 크다
         (마감일 있는 주문은 하드 제약으로 항상 필요수량을 채우므로 0).
      3) 제품군별 총 생산량 + 마스크 환산 총 생산량(마스크+마스크_멀티시트+
         (용기+튜브)*MASK_EQUIVALENT_WEIGHT) + 총 인원/인원당 환산생산량/
         총 노동시간/총 인건비
      4) 날짜별 실투입효율(output/labor_utilization.py와 동일한 정의:
         그날 필요인원시간 / 가용인원시간)

    result.is_feasible이 False면(실행 가능한 스케줄이 없으면) 만들
    내용이 없으므로 생성을 생략한다.
    """
    if not result.is_feasible:
        print("[정보] 실행 가능한 스케줄이 없어 plan_report.xlsx 생성을 생략합니다.")
        return

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def day_to_date(day):
        return None if day is None else reference_date + timedelta(days=day - 1)

    # ---- 사전 집계 ----
    workers_lookup = _build_workers_lookup(lines, orders)

    first_produce_day: dict[str, int] = {}
    required_person_hours_by_day: dict[int, float] = {}
    for line_id, entries in result.line_activity.items():
        for day, _slot_label, activity, oid in entries:
            # activity는 "produce:<product_id>" / "setup:<product_id>" / "idle" 형태
            # (scheduling/report.py의 save_outputs/plot_gantt과 동일한 파싱).
            kind, _, _product = activity.partition(":")
            if kind != "produce" or not oid:
                continue
            if oid not in first_produce_day or day < first_produce_day[oid]:
                first_produce_day[oid] = day
            workers = workers_lookup.get((line_id, oid), 0)
            required_person_hours_by_day[day] = required_person_hours_by_day.get(day, 0) + workers

    produced_by_category: dict[str, float] = {}
    for o in orders:
        f = result.order_fulfillment[o.order_id]
        produced_by_category[o.category] = produced_by_category.get(o.category, 0) + f["produced"]

    mask_qty = produced_by_category.get("마스크", 0)
    mask_sheet_qty = produced_by_category.get("마스크_멀티시트", 0)
    container_qty = produced_by_category.get("용기", 0)
    tube_qty = produced_by_category.get("튜브", 0)
    mask_equivalent_total = mask_qty + mask_sheet_qty + (container_qty + tube_qty) * MASK_EQUIVALENT_WEIGHT

    total_workforce = sum(result.daily_workforce.values())
    per_person_output = mask_equivalent_total / total_workforce if total_workforce else 0.0

    regular_hours_per_day = SLOTS_PER_DAY - len(OVERTIME_LOCAL_SLOTS)
    total_labor_hours = sum(
        result.daily_workforce[d] * regular_hours_per_day
        + result.overtime_workers[d]["17-18"] + result.overtime_workers[d]["18-19"]
        for d in result.daily_workforce
    )

    # ---- 스타일 ----
    TITLE_FONT = Font(bold=True, size=14)
    SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor="2F5597")
    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
    BOLD = Font(bold=True)
    NUM_FMT = "#,##0"
    DATE_FMT = "yyyy-mm-dd"
    PCT_FMT = "0.0%"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "생산계획 리포트"

    row = 1

    def put(r, c, value, *, font=None, fill=None, number_format=None, align=None):
        cell = ws.cell(row=r, column=c, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if number_format:
            cell.number_format = number_format
        if align:
            cell.alignment = Alignment(horizontal=align)
        return cell

    def section_title(r, text, span=8):
        put(r, 1, text, font=SECTION_FONT, fill=SECTION_FILL)
        for c in range(2, span + 1):
            put(r, c, None, fill=SECTION_FILL)
        return r + 1

    def table_header(r, headers):
        for c, h in enumerate(headers, start=1):
            put(r, c, h, font=HEADER_FONT, fill=HEADER_FILL, align="center")
        return r + 1

    # ---- 1) 기준 정보 + 제외 사유별 개수 ----
    put(row, 1, "생산계획 리포트", font=TITLE_FONT)
    row += 1
    put(row, 1, f"기준일: {reference_date.isoformat()}   계획기간: {config.horizon_days}일   "
                f"생성시각: {date.today().isoformat()}")
    row += 2

    row = section_title(row, "주문 필터링 요약", span=2)
    row = table_header(row, ["구분", "건수"])
    exclusion_rows = [
        ("스캔한 행 수", load_stats["scanned"]),
        ("완료/취소 상태 제외", load_stats["excluded_status"]),
        ("납기 해석 불가 제외", load_stats["excluded_deadline_unresolved"]),
        ("납기 지남 제외", load_stats["excluded_deadline_passed"]),
        ("잔량 0 이하 제외(원본)", load_stats["excluded_nonpositive_qty"]),
        ("제품군/호환라인 없음 제외", filter_stats["excluded_category"]),
        (f"납기 {deadline_window_days}일 초과 제외", filter_stats["excluded_deadline"]),
        ("잔량 0 이하 제외(ramp 반영 후)", filter_stats["excluded_nonpositive_qty"]),
    ]
    for label, count in exclusion_rows:
        put(row, 1, label)
        put(row, 2, count, number_format=NUM_FMT)
        row += 1
    put(row, 1, "최종 포함(계획 대상)", font=BOLD)
    put(row, 2, filter_stats["included"], font=BOLD, number_format=NUM_FMT)
    row += 3

    # ---- 2) 주문별 상세 ----
    row = section_title(row, "주문별 상세", span=8)
    row = table_header(row, ["주문ID", "발주처", "제품명", "납기일", "생산량", "생산잔량", "최초생산일", "생산완료일"])
    for o in sorted(orders, key=lambda o: (o.deadline_day is None, o.deadline_day or 0)):
        f = result.order_fulfillment[o.order_id]
        backlog = f["final_backlog"] if f["final_backlog"] else 0
        put(row, 1, o.order_id)
        put(row, 2, o.vendor)
        put(row, 3, o.product_name or o.product_id)
        put(row, 4, "ASAP" if o.deadline_day is None else day_to_date(o.deadline_day), number_format=DATE_FMT)
        put(row, 5, f["produced"], number_format=NUM_FMT)
        put(row, 6, backlog, number_format=NUM_FMT)
        put(row, 7, day_to_date(first_produce_day.get(o.order_id)), number_format=DATE_FMT)
        put(row, 8, day_to_date(f["completion_day"]), number_format=DATE_FMT)
        row += 1
    row += 1

    # ---- 3) 제품군별 총 생산량 + 종합 지표 ----
    row = section_title(row, "제품군별 총 생산량", span=2)
    row = table_header(row, ["제품군", "총 생산량"])
    for category, qty in sorted(produced_by_category.items(), key=lambda kv: -kv[1]):
        put(row, 1, category)
        put(row, 2, qty, number_format=NUM_FMT)
        row += 1
    row += 1

    row = section_title(row, "종합 지표", span=2)
    summary_rows = [
        (f"마스크 환산 총 생산량 (마스크+멀티시트+(용기+튜브)x{MASK_EQUIVALENT_WEIGHT})", mask_equivalent_total, NUM_FMT),
        ("총 인원 (인원-일)", total_workforce, NUM_FMT),
        ("인원당 환산 생산량", round(per_person_output, 1), NUM_FMT),
        ("총 노동시간 (인원-시간)", total_labor_hours, NUM_FMT),
        ("총 인건비 (원)", round(result.labor_cost or 0), NUM_FMT),
    ]
    for label, value, fmt in summary_rows:
        put(row, 1, label)
        put(row, 2, value, number_format=fmt)
        row += 1
    row += 2

    # ---- 4) 날짜별 실투입효율 ----
    row = section_title(row, "날짜별 실투입효율", span=6)
    row = table_header(row, ["일차", "날짜", "고용인원", "필요인원시간", "가용인원시간", "실투입효율"])
    for d in sorted(result.daily_workforce):
        workforce = result.daily_workforce[d]
        overtime = result.overtime_workers[d]["17-18"] + result.overtime_workers[d]["18-19"]
        available = workforce * regular_hours_per_day + overtime
        required = required_person_hours_by_day.get(d, 0)
        ratio = required / available if available else None
        put(row, 1, d, number_format=NUM_FMT)
        put(row, 2, day_to_date(d), number_format=DATE_FMT)
        put(row, 3, workforce, number_format=NUM_FMT)
        put(row, 4, required, number_format=NUM_FMT)
        put(row, 5, available, number_format=NUM_FMT)
        put(row, 6, ratio, number_format=PCT_FMT)
        row += 1

    # ---- 열 너비 ----
    widths = [22, 16, 22, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    try:
        wb.save(output_path)
    except PermissionError:
        # 가장 흔한 원인: 방금 결과를 확인하려고 엑셀에서 그 파일을 열어둔
        # 채로 다시 돌린 경우 - Windows가 파일을 잠가서 덮어쓰기가
        # 막힌다. 코드 문제가 아니므로 트레이스백 대신 원인/해결법을
        # 바로 알려주고, 나머지 산출물(CSV/PNG)은 이미 저장된 뒤라 굳이
        # 전체 실행을 실패로 만들지 않는다.
        print(
            f"[경고] {output_path} 저장 실패 - 파일이 다른 프로그램(엑셀 등)에서 열려있는 것 같습니다. "
            f"파일을 닫고 다시 실행하세요."
        )
        return
    print(f"[저장 완료] {output_path}")
